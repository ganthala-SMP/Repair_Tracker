import base64
import csv
import io
import os
import secrets
import zipfile
import json
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from functools import wraps
from pathlib import Path
from dotenv import load_dotenv

import qrcode
from reportlab.graphics.barcode import createBarcodeDrawing
from flask import Flask, render_template, redirect, url_for, request, flash, abort, g, Response, session
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, func, and_, text as sql_text, inspect as sa_inspect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename


load_dotenv(Path(__file__).resolve().parent.parent / '.env')

db = SQLAlchemy()
login_manager = LoginManager()
login_manager.login_view = 'login'

ROLE_LABELS = ['Super Admin', 'Company Admin', 'Staff', 'Technician']
COMPANY_ROLE_LABELS = ['Company Admin', 'Staff', 'Technician']
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'svg'}


def create_app():
    app = Flask(__name__)
    database_url = (os.getenv('DATABASE_URL') or 'sqlite:///repair_tracker_multi_company.db').strip()
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)

    app.config['SECRET_KEY'] = os.getenv('SECRET_KEY') or secrets.token_hex(32)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['REMEMBER_COOKIE_HTTPONLY'] = True

    Path(app.root_path, UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)

    db.init_app(app)
    login_manager.init_app(app)

    from .models import (
        AppSetting, Company, User, Customer, Part, RepairJob, RepairUpdate, Payment,
        RepairPartUsed, StatusOption, PaymentMethod, InternalNote, AuditLog, BuyingTicket, BuyingPayment,
    )
    from .translations import translations, status_key_map

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    def seed_super_admin():
        if not User.query.filter_by(role='Super Admin').first():
            admin = User(
                name='StarMobileParts Owner',
                email='info@starmobileparts.be',
                password_hash=generate_password_hash('SMP@2000'),
                role='Super Admin',
                is_active=True,
                unlimited_access=True,
                company_id=None,
            )
            db.session.add(admin)
            db.session.commit()

    def seed_statuses():
        defaults = ['Received', 'Checking', 'Waiting for Parts', 'In Repair', 'Ready', 'Delivered', 'Cancelled']
        for idx, name in enumerate(defaults, start=1):
            if not StatusOption.query.filter_by(name=name).first():
                db.session.add(StatusOption(name=name, sort_order=idx, is_active=True))
        db.session.commit()

    def seed_payment_methods():
        defaults = ['Cash', 'Card', 'Bank Transfer']
        for idx, name in enumerate(defaults, start=1):
            if not PaymentMethod.query.filter_by(name=name).first():
                db.session.add(PaymentMethod(name=name, sort_order=idx, is_active=True))
        db.session.commit()

    def seed_app_settings():
        settings = AppSetting.query.first()
        if not settings:
            db.session.add(AppSetting())
            db.session.commit()
        else:
            if not getattr(settings, 'platform_language', None):
                settings.platform_language = 'en'
                db.session.commit()

    
    def ensure_runtime_schema():
        # Lightweight runtime migration for BuyingTicket + AppSetting additions
        inspector = sa_inspect(db.engine)
        table_names = set(inspector.get_table_names())

        if 'repair_job' in table_names:
            repair_cols = {col['name'] for col in inspector.get_columns('repair_job')}
            if 'items_json' not in repair_cols:
                db.session.execute(sql_text("ALTER TABLE repair_job ADD COLUMN items_json TEXT"))
                db.session.commit()

        if 'buying_ticket' in table_names:
            buying_cols = {col['name'] for col in inspector.get_columns('buying_ticket')}
            if 'battery_percentage' not in buying_cols:
                db.session.execute(sql_text("ALTER TABLE buying_ticket ADD COLUMN battery_percentage VARCHAR(50)"))
                db.session.commit()
            if 'storage_gb' not in buying_cols:
                db.session.execute(sql_text("ALTER TABLE buying_ticket ADD COLUMN storage_gb VARCHAR(20)"))
                db.session.commit()
            if 'items_json' not in buying_cols:
                db.session.execute(sql_text("ALTER TABLE buying_ticket ADD COLUMN items_json TEXT"))
                db.session.commit()

        if 'app_setting' in table_names:
            app_cols = {col['name'] for col in inspector.get_columns('app_setting')}
            if 'menu_buying' not in app_cols:
                db.session.execute(sql_text("ALTER TABLE app_setting ADD COLUMN menu_buying VARCHAR(80)"))
                db.session.commit()
            if 'gsm_bought_status_label' not in app_cols:
                db.session.execute(sql_text("ALTER TABLE app_setting ADD COLUMN gsm_bought_status_label VARCHAR(80)"))
                db.session.commit()
            if 'gsm_sold_status_label' not in app_cols:
                db.session.execute(sql_text("ALTER TABLE app_setting ADD COLUMN gsm_sold_status_label VARCHAR(80)"))
                db.session.commit()
            db.session.execute(sql_text("UPDATE app_setting SET menu_buying = 'GSM' WHERE menu_buying IS NULL OR TRIM(menu_buying) = '' OR menu_buying = 'Buying'"))
            db.session.execute(sql_text("UPDATE app_setting SET gsm_bought_status_label = 'BOUGHT BY SMP' WHERE gsm_bought_status_label IS NULL OR TRIM(gsm_bought_status_label) = '' OR gsm_bought_status_label = 'Buying'"))
            db.session.execute(sql_text("UPDATE app_setting SET gsm_sold_status_label = 'SOLD BY SMP' WHERE gsm_sold_status_label IS NULL OR TRIM(gsm_sold_status_label) = '' OR gsm_sold_status_label = 'Selling'"))
            db.session.commit()
        if 'company' in table_names:
            company_cols = {col['name'] for col in inspector.get_columns('company')}
            company_additions = [
                ('gsm_invoice_title', "ALTER TABLE company ADD COLUMN gsm_invoice_title VARCHAR(120)"),
                ('gsm_terms_en', "ALTER TABLE company ADD COLUMN gsm_terms_en TEXT"),
                ('gsm_terms_fr', "ALTER TABLE company ADD COLUMN gsm_terms_fr TEXT"),
                ('gsm_terms_nl', "ALTER TABLE company ADD COLUMN gsm_terms_nl TEXT"),
                ('gsm_notice_en', "ALTER TABLE company ADD COLUMN gsm_notice_en TEXT"),
                ('gsm_notice_fr', "ALTER TABLE company ADD COLUMN gsm_notice_fr TEXT"),
                ('gsm_notice_nl', "ALTER TABLE company ADD COLUMN gsm_notice_nl TEXT"),
                ('print_80mm_width_mm', "ALTER TABLE company ADD COLUMN print_80mm_width_mm FLOAT"),
                ('print_80mm_padding_mm', "ALTER TABLE company ADD COLUMN print_80mm_padding_mm FLOAT"),
                ('print_80mm_header_font_px', "ALTER TABLE company ADD COLUMN print_80mm_header_font_px INTEGER"),
                ('print_80mm_body_font_px', "ALTER TABLE company ADD COLUMN print_80mm_body_font_px INTEGER"),
                ('print_80mm_notice_font_px', "ALTER TABLE company ADD COLUMN print_80mm_notice_font_px INTEGER"),
                ('print_80mm_content_shift_mm', "ALTER TABLE company ADD COLUMN print_80mm_content_shift_mm FLOAT"),
            ]
            for col_name, stmt in company_additions:
                if col_name not in company_cols:
                    db.session.execute(sql_text(stmt))
                    db.session.commit()

    with app.app_context():

            db.create_all()
            ensure_runtime_schema()
            seed_super_admin()
            seed_statuses()
            seed_payment_methods()
            seed_app_settings()

    def get_gsm_statuses():
        settings = AppSetting.query.first()
        bought = (settings.gsm_bought_status_label if settings and getattr(settings, 'gsm_bought_status_label', None) else 'BOUGHT BY SMP')
        sold = (settings.gsm_sold_status_label if settings and getattr(settings, 'gsm_sold_status_label', None) else 'SOLD BY SMP')
        return [bought, sold]

    def current_language():
        settings = AppSetting.query.first()
        platform_lang = (settings.platform_language if settings and settings.platform_language else 'en')
        if current_user.is_authenticated:
            if current_user.is_super_admin:
                return platform_lang
            if current_user.company and current_user.company.default_language:
                return current_user.company.default_language
        return platform_lang

    def t(key):
        lang = getattr(g, 'lang', current_language())
        return translations.get(lang, translations['en']).get(key, translations['en'].get(key, key))

    def status_label(status):
        key = status_key_map.get(status)
        return t(key) if key else status

    def get_statuses(active_only=True):
        q = StatusOption.query.order_by(StatusOption.sort_order.asc(), StatusOption.name.asc())
        if active_only:
            q = q.filter_by(is_active=True)
        return [row.name for row in q.all()]

    def get_payment_methods(active_only=True):
        q = PaymentMethod.query.order_by(PaymentMethod.sort_order.asc(), PaymentMethod.name.asc())
        if active_only:
            q = q.filter_by(is_active=True)
        return [row.name for row in q.all()]

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def save_logo(file_storage):
        if not file_storage or not file_storage.filename:
            return None, None
        if not allowed_file(file_storage.filename):
            return None, 'Unsupported logo file type. Use PNG, JPG, JPEG, WEBP, or SVG.'
        allowed_mimetypes = {'image/png', 'image/jpeg', 'image/webp', 'image/svg+xml'}
        if getattr(file_storage, 'mimetype', None) and file_storage.mimetype not in allowed_mimetypes:
            return None, 'Unsupported uploaded file content type.'
        ext = file_storage.filename.rsplit('.', 1)[1].lower()
        filename = secure_filename(file_storage.filename.rsplit('.', 1)[0])[:50] or 'logo'
        final_name = f'{filename}-{secrets.token_hex(6)}.{ext}'
        abs_path = Path(app.root_path, UPLOAD_FOLDER, final_name)
        try:
            file_storage.save(abs_path)
        except Exception:
            return None, 'Could not save the uploaded logo.'
        if abs_path.stat().st_size > 4 * 1024 * 1024:
            abs_path.unlink(missing_ok=True)
            return None, 'Logo file is too large.'
        return f'uploads/{final_name}', None

    def parse_date_field(value):
        value = (value or '').strip()
        if not value:
            return None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None


    def combine_phone(phone_code, phone_local):
        phone_code = (phone_code or '').strip()
        phone_local = ''.join(ch for ch in (phone_local or '').strip() if ch.isdigit())
        if phone_code == 'other':
            return phone_local
        if phone_code and phone_local:
            return f"{phone_code}{phone_local}"
        return phone_local


    def parse_gsm_items_from_form(form):
        brands = form.getlist('item_brand[]')
        models = form.getlist('item_model[]')
        gbs = form.getlist('item_gb[]')
        imeis = form.getlist('item_imei[]')
        batteries = form.getlist('item_battery[]')
        serials = form.getlist('item_serial[]')
        qtys = form.getlist('item_qty[]')
        amounts = form.getlist('item_amount[]')
        items = []
        max_len = max([len(brands), len(models), len(gbs), len(imeis), len(batteries), len(serials), len(qtys), len(amounts)] + [0])
        for i in range(max_len):
            item = {
                'brand': (brands[i] if i < len(brands) else '').strip(),
                'model': (models[i] if i < len(models) else '').strip(),
                'gb': (gbs[i] if i < len(gbs) else '').strip(),
                'imei': (imeis[i] if i < len(imeis) else '').strip(),
                'battery': (batteries[i] if i < len(batteries) else '').strip(),
                'serial': (serials[i] if i < len(serials) else '').strip(),
                'qty': int(float((qtys[i] if i < len(qtys) else '1') or 1)),
                'amount': safe_float(amounts[i] if i < len(amounts) else 0, 0),
            }
            if any([item['brand'], item['model'], item['gb'], item['imei'], item['battery'], item['serial'], item['amount']]):
                if item['qty'] <= 0:
                    item['qty'] = 1
                items.append(item)
        return items

    def ticket_items(ticket):
        raw = getattr(ticket, 'items_json', None)
        if raw:
            try:
                items = json.loads(raw)
                if isinstance(items, list) and items:
                    normalized = []
                    for item in items:
                        normalized.append({
                            'brand': (item.get('brand') or '').strip(),
                            'model': (item.get('model') or '').strip(),
                            'gb': (item.get('gb') or '').strip(),
                            'imei': (item.get('imei') or '').strip(),
                            'battery': (str(item.get('battery') or '')).strip(),
                            'serial': (item.get('serial') or '').strip(),
                            'qty': int(float(item.get('qty') or 1)),
                            'amount': safe_float(item.get('amount'), 0),
                        })
                    return normalized
            except Exception:
                pass
        return [{
            'brand': getattr(ticket, 'device_brand', '') or '',
            'model': getattr(ticket, 'device_model', '') or '',
            'gb': getattr(ticket, 'storage_gb', '') or '',
            'imei': getattr(ticket, 'imei', '') or '',
            'battery': getattr(ticket, 'battery_percentage', '') or '',
            'serial': getattr(ticket, 'serial_number', '') or '',
            'qty': 1,
            'amount': safe_float(getattr(ticket, 'estimated_cost', 0), 0),
        }]

    def format_gsm_description(item):
        parts = []
        for key in ('brand', 'model', 'gb'):
            val = (item.get(key) or '').strip()
            if val:
                parts.append(val)
        return ' '.join(parts).strip()

    def ticket_total(ticket):
        return sum((item.get('qty') or 1) * safe_float(item.get('amount'), 0) for item in ticket_items(ticket))

    def repair_items(repair):
        raw = getattr(repair, 'items_json', None)
        if raw:
            try:
                items = json.loads(raw)
                if isinstance(items, list) and items:
                    normalized = []
                    for item in items:
                        normalized.append({
                            'brand': (item.get('brand') or '').strip(),
                            'model': (item.get('model') or '').strip(),
                            'imei': (item.get('imei') or '').strip(),
                            'serial': (item.get('serial') or '').strip(),
                            'issue': (item.get('issue') or '').strip(),
                            'amount': safe_float(item.get('amount'), 0),
                        })
                    return normalized
            except Exception:
                pass
        return [{
            'brand': getattr(repair, 'device_brand', '') or '',
            'model': getattr(repair, 'device_model', '') or '',
            'imei': getattr(repair, 'imei', '') or '',
            'serial': getattr(repair, 'serial_number', '') or '',
            'issue': getattr(repair, 'issue_description', '') or '',
            'amount': safe_float(getattr(repair, 'estimated_cost', 0), 0),
        }]

    def format_repair_description(item):
        device = ' '.join([p for p in [(item.get('brand') or '').strip(), (item.get('model') or '').strip()] if p]).strip()
        issue = (item.get('issue') or '').strip()
        if device and issue:
            return f"{device} - {issue}"
        return device or issue

    def repair_items_total(repair):
        return sum(safe_float(item.get('amount'), 0) for item in repair_items(repair))

    def parse_repair_items_from_form(form):
        brands = form.getlist('repair_item_brand[]')
        models = form.getlist('repair_item_model[]')
        imeis = form.getlist('repair_item_imei[]')
        serials = form.getlist('repair_item_serial[]')
        issues = form.getlist('repair_item_issue[]')
        amounts = form.getlist('repair_item_amount[]')
        items = []
        max_len = max([len(brands), len(models), len(imeis), len(serials), len(issues), len(amounts)] + [0])
        for i in range(max_len):
            item = {
                'brand': (brands[i] if i < len(brands) else '').strip(),
                'model': (models[i] if i < len(models) else '').strip(),
                'imei': (imeis[i] if i < len(imeis) else '').strip(),
                'serial': (serials[i] if i < len(serials) else '').strip(),
                'issue': (issues[i] if i < len(issues) else '').strip(),
                'amount': safe_float(amounts[i] if i < len(amounts) else 0, 0),
            }
            if item['brand'] or item['model'] or item['imei'] or item['serial'] or item['issue'] or item['amount']:
                items.append(item)
        return items

    def company_access_allowed(company):
        if not company:
            return False, 'Company account is missing.'
        if not company.is_active:
            return False, 'This company account is disabled.'
        if not company.unlimited_access and company.active_until and company.active_until < date.today():
            return False, 'This company subscription has expired.'
        return True, ''

    def is_account_allowed(user):
        if not user.is_active:
            return False, 'This user account is disabled.'
        if user.company:
            ok, reason = company_access_allowed(user.company)
            if not ok:
                return False, reason
        if not user.unlimited_access and user.active_until and user.active_until < date.today():
            return False, 'This user account has expired.'
        return True, ''

    def company_required(fn):
        @wraps(fn)
        @login_required
        def wrapper(*args, **kwargs):
            if current_user.is_super_admin:
                flash('This page is for company accounts.', 'warning')
                return redirect(url_for('super_admin_dashboard'))
            allowed, reason = is_account_allowed(current_user)
            if not current_user.company_id or not current_user.company or not current_user.company.is_active or not allowed:
                logout_user()
                flash(reason or 'Your company is disabled or missing.', 'danger')
                return redirect(url_for('login'))
            auto_archive_company_repairs(current_user.company_id)
            return fn(*args, **kwargs)
        return wrapper

    def roles_required(*roles):
        def decorator(fn):
            @wraps(fn)
            @login_required
            def wrapper(*args, **kwargs):
                if current_user.role not in roles:
                    flash('You do not have permission for that page.', 'danger')
                    if current_user.is_super_admin:
                        return redirect(url_for('super_admin_dashboard'))
                    return redirect(url_for('dashboard'))
                return fn(*args, **kwargs)
            return wrapper
        return decorator

    def company_query(model):
        if current_user.is_super_admin:
            abort(403)
        if hasattr(model, 'is_archived'):
            return model.query.filter_by(company_id=current_user.company_id, is_archived=False)
        return model.query.filter_by(company_id=current_user.company_id)

    def get_company_or_404(company_id):
        return Company.query.get_or_404(company_id)

    def company_object_or_404(model, object_id):
        obj = model.query.get_or_404(object_id)
        if current_user.is_super_admin:
            return obj
        if getattr(obj, 'company_id', None) != current_user.company_id:
            abort(404)
        return obj

    def _next_company_sequence(company_id, field_name='repair_code'):
        prefix = f'REP-{company_id:02d}-'
        max_num = 0
        rows = RepairJob.query.filter_by(company_id=company_id).all()
        for row in rows:
            value = getattr(row, field_name, '') or ''
            if value.startswith(prefix):
                try:
                    num = int(value.split('-')[-1])
                    if num > max_num:
                        max_num = num
                except ValueError:
                    continue
        return max_num + 1

    def generate_repair_code(company_id):
        next_num = _next_company_sequence(company_id, 'repair_code')
        return f'REP-{company_id:02d}-{next_num:04d}'

    def generate_invoice_number(company_id):
        next_num = _next_company_sequence(company_id, 'invoice_number')
        return f'REP-{company_id:02d}-{next_num:04d}'

    def _next_buying_sequence(company_id):
        prefix = f'TICK-{company_id:02d}-'
        max_num = 0
        rows = BuyingTicket.query.filter_by(company_id=company_id).all()
        for row in rows:
            value = row.ticket_code or ''
            if value.startswith(prefix):
                try:
                    num = int(value.split('-')[-1])
                    if num > max_num:
                        max_num = num
                except ValueError:
                    continue
        return max_num + 1

    def generate_ticket_code(company_id):
        next_num = _next_buying_sequence(company_id)
        return f'TICK-{company_id:02d}-{next_num:04d}'

    def generate_buying_invoice_number(company_id):
        next_num = _next_buying_sequence(company_id)
        return f'TICK-{company_id:02d}-{next_num:04d}'


    def qr_data_uri(data: str):
        img = qrcode.make(data)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        encoded = base64.b64encode(buf.getvalue()).decode('ascii')
        return f'data:image/png;base64,{encoded}'

    def barcode_svg_uri(data: str):
        drawing = createBarcodeDrawing('Code128', value=data, barHeight=28, humanReadable=True)
        svg = drawing.asString('svg')
        if isinstance(svg, str):
            svg = svg.encode('utf-8')
        encoded = base64.b64encode(svg).decode('ascii')
        return f'data:image/svg+xml;base64,{encoded}'

    def log_audit(action, target_type=None, target_id=None, detail=''):
        if not current_user.is_authenticated:
            return
        entry = AuditLog(
            company_id=None if current_user.is_super_admin else current_user.company_id,
            user_id=current_user.id,
            action=action,
            target_type=target_type,
            target_id=target_id,
            detail=detail[:2000] if detail else ''
        )
        db.session.add(entry)

    def get_default_country_code():
        raw = (os.getenv('DEFAULT_COUNTRY_CODE') or '+32').strip()
        if not raw:
            raw = '+32'
        if raw.startswith('00'):
            raw = '+' + raw[2:]
        if not raw.startswith('+'):
            raw = '+' + raw
        digits = ''.join(ch for ch in raw if ch.isdigit())
        return f'+{digits}' if digits else '+32'

    def normalize_phone(phone, default_country_code=None):
        phone = (phone or '').strip()
        if not phone:
            return ''
        phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if phone.startswith('+'):
            return '+' + ''.join(ch for ch in phone[1:] if ch.isdigit())
        if phone.startswith('00'):
            return '+' + ''.join(ch for ch in phone[2:] if ch.isdigit())
        digits = ''.join(ch for ch in phone if ch.isdigit())
        if not digits:
            return ''
        default_country_code = default_country_code or get_default_country_code()
        if digits.startswith('0'):
            digits = digits[1:]
        return f"{default_country_code}{digits}"

    def build_customer_whatsapp_url(repair):
        customer_phone = normalize_phone(repair.customer.phone)
        if not customer_phone:
            return None
        company_name = repair.company.name if repair.company else 'Our repair shop'
        status = repair.status or 'updated'
        if status in ('Ready', 'Delivered'):
            message = f"Hello {repair.customer.name}, your repair {repair.repair_code} for {repair.device_brand} {repair.device_model} is {status.lower()} at {company_name}. Thank you."
        else:
            message = f"Hello {repair.customer.name}, here is an update for your repair {repair.repair_code} at {company_name}: status is {status}."
        from urllib.parse import quote
        return f"https://wa.me/{customer_phone.replace('+','')}?text={quote(message)}"

    def auto_archive_company_repairs(company_id):
        threshold = datetime.utcnow() - timedelta(days=120)
        old_jobs = RepairJob.query.filter(
            RepairJob.company_id == company_id,
            RepairJob.is_archived.is_(False),
            RepairJob.status.in_(['Delivered', 'Cancelled']),
            RepairJob.created_at < threshold,
        ).all()
        changed = False
        for job in old_jobs:
            job.is_archived = True
            job.archive_reason = 'Auto archived'
            changed = True
        if changed:
            db.session.commit()



    def safe_int(value, default=0):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def safe_float(value, default=0.0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def csv_template_response(filename, headers):
        sio = io.StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        return Response(sio.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename}'})

    def import_customers_csv(file_storage, company_id):
        if not file_storage or not file_storage.filename:
            return 0, 0, ['Please choose a CSV or Excel file.']
        lower = file_storage.filename.lower()
        if lower.endswith('.csv'):
            try:
                raw = file_storage.stream.read().decode('utf-8-sig')
            except Exception:
                return 0, 0, ['Could not read CSV file. Please use UTF-8 CSV.']
            reader = csv.DictReader(io.StringIO(raw))
        elif lower.endswith('.xlsx'):
            try:
                wb = load_workbook(filename=io.BytesIO(file_storage.read()), data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [x if x is not None else '' for x in row])))
                reader = rows
            except Exception:
                return 0, 0, ['Could not read Excel file. Please use .xlsx format.']
        else:
            return 0, 0, ['Please upload a CSV or Excel file.']
        required = {'name', 'phone'}
        fieldnames = set(h.strip() for h in (reader.fieldnames if hasattr(reader, 'fieldnames') and reader.fieldnames else (reader[0].keys() if reader else [])))
        if not fieldnames or not required.issubset(fieldnames):
            return 0, 0, ['Import file must include at least: name, phone']
        created = 0
        skipped = 0
        errors = []
        for idx, row in enumerate(reader, start=2):
            name = (row.get('name') or '').strip()
            phone = (row.get('phone') or '').strip()
            email = (row.get('email') or '').strip()
            if not name or not phone:
                errors.append(f'Row {idx}: name and phone are required.')
                continue
            exists = Customer.query.filter_by(company_id=company_id, name=name, phone=phone).first()
            if exists:
                skipped += 1
                continue
            db.session.add(Customer(company_id=company_id, name=name, phone=phone, email=email))
            created += 1
        db.session.commit()
        return created, skipped, errors

    def import_parts_csv(file_storage, company_id):
        if not file_storage or not file_storage.filename:
            return 0, 0, ['Please choose a CSV or Excel file.']
        lower = file_storage.filename.lower()
        if lower.endswith('.csv'):
            try:
                raw = file_storage.stream.read().decode('utf-8-sig')
            except Exception:
                return 0, 0, ['Could not read CSV file. Please use UTF-8 CSV.']
            reader = csv.DictReader(io.StringIO(raw))
        elif lower.endswith('.xlsx'):
            try:
                wb = load_workbook(filename=io.BytesIO(file_storage.read()), data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [x if x is not None else '' for x in row])))
                reader = rows
            except Exception:
                return 0, 0, ['Could not read Excel file. Please use .xlsx format.']
        else:
            return 0, 0, ['Please upload a CSV or Excel file.']
        required = {'name'}
        fieldnames = set(h.strip() for h in (reader.fieldnames if hasattr(reader, 'fieldnames') and reader.fieldnames else (reader[0].keys() if reader else [])))
        if not fieldnames or not required.issubset(fieldnames):
            return 0, 0, ['Import file must include at least: name']
        created = 0
        skipped = 0
        errors = []
        for idx, row in enumerate(reader, start=2):
            name = (row.get('name') or '').strip()
            compatible_model = (row.get('compatible_model') or '').strip()
            if not name:
                errors.append(f'Row {idx}: name is required.')
                continue
            quantity = safe_int(row.get('quantity'), 0)
            selling_price = safe_float(row.get('selling_price'), 0.0)
            low_stock_limit = safe_int(row.get('low_stock_limit'), 0)
            exists = Part.query.filter_by(company_id=company_id, name=name, compatible_model=compatible_model).first()
            if exists:
                skipped += 1
                continue
            db.session.add(Part(company_id=company_id, name=name, compatible_model=compatible_model, quantity=quantity, selling_price=selling_price, low_stock_limit=low_stock_limit))
            created += 1
        db.session.commit()
        return created, skipped, errors


    def export_company_zip(company):
        mem = io.BytesIO()
        with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
            datasets = {
                'customers.csv': company.customers,
                'parts.csv': company.parts,
                'repairs.csv': [r for r in company.repair_jobs],
                'payments.csv': company.payments,
            }
            for filename, rows in datasets.items():
                csv_io = io.StringIO()
                writer = csv.writer(csv_io)
                if filename == 'customers.csv':
                    writer.writerow(['id', 'name', 'phone', 'email', 'created_at'])
                    for row in rows:
                        writer.writerow([row.id, row.name, row.phone, row.email or '', row.created_at])
                elif filename == 'parts.csv':
                    writer.writerow(['id', 'name', 'compatible_model', 'quantity', 'selling_price', 'low_stock_limit'])
                    for row in rows:
                        writer.writerow([row.id, row.name, row.compatible_model or '', row.quantity, row.selling_price, row.low_stock_limit])
                elif filename == 'repairs.csv':
                    writer.writerow(['id', 'repair_code', 'invoice_number', 'customer', 'phone', 'brand', 'model', 'serial', 'imei', 'status', 'estimated_cost', 'deposit_amount', 'due_date', 'created_at'])
                    for row in rows:
                        writer.writerow([row.id, row.repair_code, row.invoice_number, row.customer.name, row.customer.phone, row.device_brand, row.device_model, row.serial_number or '', row.imei or '', row.status, row.estimated_cost, row.deposit_amount, row.due_date or '', row.created_at])
                elif filename == 'payments.csv':
                    writer.writerow(['id', 'repair_code', 'amount', 'method', 'created_at'])
                    for row in rows:
                        writer.writerow([row.id, row.repair_job.repair_code, row.amount, row.method, row.created_at])
                zf.writestr(filename, csv_io.getvalue())
        mem.seek(0)
        return mem

    def export_company_json(company):
        data = {
            'company': {
                'name': company.name,
                'phone': company.phone or '',
                'email': company.email or '',
                'address': company.address or '',
                'vat_number': company.vat_number or '',
                'default_language': company.default_language or 'en',
                'invoice_terms_en': company.invoice_terms_en or '',
                'invoice_terms_fr': company.invoice_terms_fr or '',
                'invoice_terms_nl': company.invoice_terms_nl or '',
                'storage_notice_en': company.storage_notice_en or '',
                'storage_notice_fr': company.storage_notice_fr or '',
                'storage_notice_nl': company.storage_notice_nl or '',
                'gsm_invoice_title': company.gsm_invoice_title or '',
                'gsm_terms_en': company.gsm_terms_en or '',
                'gsm_terms_fr': company.gsm_terms_fr or '',
                'gsm_terms_nl': company.gsm_terms_nl or '',
                'gsm_notice_en': company.gsm_notice_en or '',
                'gsm_notice_fr': company.gsm_notice_fr or '',
                'gsm_notice_nl': company.gsm_notice_nl or '',
            },
            'customers': [
                {'name': c.name, 'phone': c.phone, 'email': c.email or ''}
                for c in company.customers
            ],
            'parts': [
                {
                    'name': p.name,
                    'compatible_model': p.compatible_model or '',
                    'quantity': p.quantity or 0,
                    'selling_price': p.selling_price or 0,
                    'low_stock_limit': p.low_stock_limit or 0,
                }
                for p in company.parts
            ],
            'repairs': [],
        }
        data['gsm_tickets'] = []
        for r in company.repair_jobs:
            data['repairs'].append({
                'repair_code': r.repair_code,
                'invoice_number': r.invoice_number,
                'customer_name': r.customer.name,
                'customer_phone': r.customer.phone,
                'device_brand': r.device_brand,
                'device_model': r.device_model,
                'serial_number': r.serial_number or '',
                'imei': r.imei or '',
                'issue_description': r.issue_description or '',
                'status': r.status,
                'estimated_cost': r.estimated_cost or 0,
                'deposit_amount': r.deposit_amount or 0,
                'due_date': r.due_date.isoformat() if r.due_date else '',
                'customer_note': r.customer_note or '',
                'items': repair_items(r),
                'created_at': r.created_at.isoformat() if r.created_at else '',
                'payments': [
                    {'amount': p.amount, 'method': p.method, 'created_at': p.created_at.isoformat() if p.created_at else ''}
                    for p in r.payments
                ],
                'internal_notes': [
                    {'note': n.note, 'created_at': n.created_at.isoformat() if n.created_at else '', 'user_name': n.user.name if n.user else ''}
                    for n in r.internal_notes
                ],
            })
        for bt in BuyingTicket.query.filter_by(company_id=company.id).all():
            data['gsm_tickets'].append({
                'ticket_code': bt.ticket_code,
                'invoice_number': bt.invoice_number,
                'customer_name': bt.customer.name,
                'customer_phone': bt.customer.phone,
                'device_brand': bt.device_brand,
                'device_model': bt.device_model,
                'storage_gb': bt.storage_gb or '',
                'serial_number': bt.serial_number or '',
                'imei': bt.imei or '',
                'battery_percentage': bt.battery_percentage or '',
                'note': bt.issue_description or '',
                'status': bt.status,
                'estimated_cost': bt.estimated_cost or 0,
                'due_date': bt.due_date.isoformat() if bt.due_date else '',
                'payments': [{'amount': p.amount, 'method': p.method} for p in bt.payments],
            })
        return data


    def export_company_xlsx(company):
        wb = Workbook()
        ws = wb.active
        ws.title = 'company'
        ws.append(['field', 'value'])
        for field in ['name','phone','email','address','vat_number','default_language','invoice_terms_en','invoice_terms_fr','invoice_terms_nl','storage_notice_en','storage_notice_fr','storage_notice_nl','gsm_invoice_title','gsm_terms_en','gsm_terms_fr','gsm_terms_nl','gsm_notice_en','gsm_notice_fr','gsm_notice_nl']:
            ws.append([field, getattr(company, field) or ''])

        ws = wb.create_sheet('customers')
        ws.append(['name', 'phone', 'email'])
        for c in company.customers:
            ws.append([c.name, c.phone, c.email or ''])

        ws = wb.create_sheet('parts')
        ws.append(['name', 'compatible_model', 'quantity', 'selling_price', 'low_stock_limit'])
        for p in company.parts:
            ws.append([p.name, p.compatible_model or '', p.quantity, p.selling_price, p.low_stock_limit])

        ws = wb.create_sheet('repairs')
        ws.append(['repair_code','invoice_number','customer_name','customer_phone','device_brand','device_model','serial_number','imei','issue_description','status','estimated_cost','deposit_amount','due_date','customer_note','items_json'])
        for r in company.repair_jobs:
            ws.append([r.repair_code, r.invoice_number, r.customer.name, r.customer.phone, r.device_brand, r.device_model, r.serial_number or '', r.imei or '', r.issue_description or '', r.status, r.estimated_cost or 0, r.deposit_amount or 0, r.due_date.isoformat() if r.due_date else '', r.customer_note or '', json.dumps(repair_items(r))])

        ws = wb.create_sheet('payments')
        ws.append(['repair_code', 'amount', 'method'])
        for p in company.payments:
            ws.append([p.repair_job.repair_code, p.amount, p.method])

        ws = wb.create_sheet('gsm_tickets')
        ws.append(['ticket_code','invoice_number','customer_name','customer_phone','device_brand','device_model','storage_gb','serial_number','imei','battery_percentage','note','status','estimated_cost','due_date'])
        for bt in BuyingTicket.query.filter_by(company_id=company.id).all():
            ws.append([bt.ticket_code, bt.invoice_number, bt.customer.name, bt.customer.phone, bt.device_brand, bt.device_model, bt.storage_gb or '', bt.serial_number or '', bt.imei or '', bt.battery_percentage or '', bt.issue_description or '', bt.status, bt.estimated_cost or 0, bt.due_date.isoformat() if bt.due_date else ''])

        ws = wb.create_sheet('gsm_payments')
        ws.append(['ticket_code', 'amount', 'method'])
        for p in BuyingPayment.query.filter_by(company_id=company.id).all():
            ws.append([p.buying_ticket.ticket_code, p.amount, p.method])

        ws = wb.create_sheet('internal_notes')
        ws.append(['repair_code', 'note'])
        for n in company.internal_notes:
            ws.append([n.repair_job.repair_code, n.note])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def import_company_xlsx(file_storage, company):
        if not file_storage or not file_storage.filename:
            return False, 'Please choose an Excel backup file.'
        try:
            wb = load_workbook(filename=io.BytesIO(file_storage.read()), data_only=True)
        except Exception:
            return False, 'Could not read Excel backup file.'

        repair_ids = [r.id for r in RepairJob.query.filter_by(company_id=company.id).all()]
        if repair_ids:
            RepairPartUsed.query.filter(RepairPartUsed.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            Payment.query.filter(Payment.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            RepairUpdate.query.filter(RepairUpdate.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            InternalNote.query.filter(InternalNote.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            RepairJob.query.filter(RepairJob.id.in_(repair_ids)).delete(synchronize_session=False)
        BuyingPayment.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        BuyingTicket.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        Customer.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        Part.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        db.session.flush()

        if 'company' in wb.sheetnames:
            rows = list(wb['company'].iter_rows(min_row=2, values_only=True))
            data = {k:v for k,v in rows if k}
            for field in ['name','phone','email','address','vat_number','default_language','invoice_terms_en','invoice_terms_fr','invoice_terms_nl','storage_notice_en','storage_notice_fr','storage_notice_nl','gsm_invoice_title','gsm_terms_en','gsm_terms_fr','gsm_terms_nl','gsm_notice_en','gsm_notice_fr','gsm_notice_nl']:
                if field in data and data[field] is not None:
                    setattr(company, field, data[field])
            company.invoice_terms = company.invoice_terms_en or company.invoice_terms
            company.storage_notice = company.storage_notice_en or company.storage_notice

        customer_map = {}
        if 'customers' in wb.sheetnames:
            for row in wb['customers'].iter_rows(min_row=2, values_only=True):
                name, phone, email = row[:3]
                name = (name or '').strip()
                phone = (phone or '').strip()
                if not name or not phone:
                    continue
                cust = Customer(company_id=company.id, name=name, phone=phone, email=(email or '').strip())
                db.session.add(cust)
                db.session.flush()
                customer_map[(cust.name, cust.phone)] = cust

        if 'parts' in wb.sheetnames:
            for row in wb['parts'].iter_rows(min_row=2, values_only=True):
                name, compatible_model, quantity, selling_price, low_stock_limit = (list(row) + [None]*5)[:5]
                name = (name or '').strip()
                if not name:
                    continue
                db.session.add(Part(company_id=company.id, name=name, compatible_model=(compatible_model or '').strip(), quantity=safe_int(quantity,0), selling_price=safe_float(selling_price,0), low_stock_limit=safe_int(low_stock_limit,0)))

        repair_map = {}
        if 'repairs' in wb.sheetnames:
            for row in wb['repairs'].iter_rows(min_row=2, values_only=True):
                vals = (list(row) + [None]*14)[:14]
                repair_code, invoice_number, customer_name, customer_phone, device_brand, device_model, serial_number, imei, issue_description, status, estimated_cost, deposit_amount, due_date, customer_note = vals
                cust = customer_map.get(((customer_name or '').strip(), (customer_phone or '').strip()))
                if not cust:
                    continue
                repair = RepairJob(company_id=company.id, repair_code=(repair_code or generate_repair_code(company.id)).strip(), invoice_number=(invoice_number or generate_invoice_number(company.id)).strip(), customer_id=cust.id, device_brand=(device_brand or '').strip(), device_model=(device_model or '').strip(), serial_number=(serial_number or '').strip(), imei=(imei or '').strip(), issue_description=(issue_description or '').strip(), status=(status or 'Received').strip(), estimated_cost=safe_float(estimated_cost,0), deposit_amount=safe_float(deposit_amount,0), due_date=parse_date_field(due_date), customer_note=(customer_note or '').strip(), items_json=(items_json or json.dumps([{'brand': (device_brand or '').strip(), 'model': (device_model or '').strip(), 'serial': (serial_number or '').strip(), 'imei': (imei or '').strip(), 'issue': (issue_description or '').strip(), 'amount': safe_float(estimated_cost,0)}])))
                db.session.add(repair)
                db.session.flush()
                repair_map[repair.repair_code] = repair

        if 'payments' in wb.sheetnames:
            for row in wb['payments'].iter_rows(min_row=2, values_only=True):
                repair_code, amount, method = (list(row)+[None]*3)[:3]
                repair = repair_map.get((repair_code or '').strip())
                if repair:
                    db.session.add(Payment(company_id=company.id, repair_job_id=repair.id, amount=safe_float(amount,0), method=(method or 'Cash').strip() or 'Cash'))

        gsm_map = {}
        if 'gsm_tickets' in wb.sheetnames:
            for row in wb['gsm_tickets'].iter_rows(min_row=2, values_only=True):
                vals = (list(row)+[None]*14)[:14]
                ticket_code, invoice_number, customer_name, customer_phone, device_brand, device_model, storage_gb, serial_number, imei, battery_percentage, note, status, estimated_cost, due_date = vals
                cust = customer_map.get(((customer_name or '').strip(), (customer_phone or '').strip()))
                if not cust:
                    continue
                bt = BuyingTicket(company_id=company.id, ticket_code=(ticket_code or generate_ticket_code(company.id)).strip(), invoice_number=(invoice_number or generate_buying_invoice_number(company.id)).strip(), customer_id=cust.id, device_brand=(device_brand or '').strip(), device_model=(device_model or '').strip(), storage_gb=(storage_gb or '').strip(), serial_number=(serial_number or '').strip(), imei=(imei or '').strip(), battery_percentage=(battery_percentage or '').strip(), issue_description=(note or '').strip(), status=(status or 'Buying').strip(), estimated_cost=safe_float(estimated_cost,0), deposit_amount=0, due_date=parse_date_field(due_date), customer_note='')
                db.session.add(bt); db.session.flush(); gsm_map[bt.ticket_code]=bt
        if 'gsm_payments' in wb.sheetnames:
            for row in wb['gsm_payments'].iter_rows(min_row=2, values_only=True):
                ticket_code, amount, method = (list(row)+[None]*3)[:3]
                bt = gsm_map.get((ticket_code or '').strip())
                if bt:
                    db.session.add(BuyingPayment(company_id=company.id, buying_ticket_id=bt.id, amount=safe_float(amount,0), method=(method or 'Cash').strip() or 'Cash'))
        if 'internal_notes' in wb.sheetnames:
            for row in wb['internal_notes'].iter_rows(min_row=2, values_only=True):
                repair_code, note = (list(row)+[None]*2)[:2]
                repair = repair_map.get((repair_code or '').strip())
                if repair and (note or '').strip():
                    db.session.add(InternalNote(company_id=company.id, repair_job_id=repair.id, user_id=current_user.id, note=(note or '').strip()))

        db.session.commit()
        return True, 'Full Excel backup imported successfully.'

    def import_company_json(file_storage, company):
        if not file_storage or not file_storage.filename:
            return False, 'Please choose a backup JSON file.'
        try:
            payload = json.loads(file_storage.stream.read().decode('utf-8-sig'))
        except Exception:
            return False, 'Could not read JSON backup file.'
        if not isinstance(payload, dict):
            return False, 'Invalid backup format.'

        # wipe current company operational data
        repair_ids = [r.id for r in RepairJob.query.filter_by(company_id=company.id).all()]
        if repair_ids:
            RepairPartUsed.query.filter(RepairPartUsed.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            Payment.query.filter(Payment.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            RepairUpdate.query.filter(RepairUpdate.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            InternalNote.query.filter(InternalNote.repair_job_id.in_(repair_ids)).delete(synchronize_session=False)
            RepairJob.query.filter(RepairJob.id.in_(repair_ids)).delete(synchronize_session=False)
        BuyingPayment.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        BuyingTicket.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        Customer.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        Part.query.filter_by(company_id=company.id).delete(synchronize_session=False)
        db.session.flush()

        company_data = payload.get('company', {}) or {}
        for field in ['name','phone','email','address','vat_number','default_language','invoice_terms_en','invoice_terms_fr','invoice_terms_nl','storage_notice_en','storage_notice_fr','storage_notice_nl','gsm_invoice_title','gsm_terms_en','gsm_terms_fr','gsm_terms_nl','gsm_notice_en','gsm_notice_fr','gsm_notice_nl']:
            if field in company_data:
                setattr(company, field, company_data.get(field) or getattr(company, field))
        company.invoice_terms = company.invoice_terms_en or company.invoice_terms
        company.storage_notice = company.storage_notice_en or company.storage_notice

        customer_map = {}
        for c in payload.get('customers', []):
            cust = Customer(company_id=company.id, name=(c.get('name') or '').strip(), phone=(c.get('phone') or '').strip(), email=(c.get('email') or '').strip())
            if not cust.name or not cust.phone:
                continue
            db.session.add(cust)
            db.session.flush()
            customer_map[(cust.name, cust.phone)] = cust

        for p in payload.get('parts', []):
            part = Part(
                company_id=company.id,
                name=(p.get('name') or '').strip(),
                compatible_model=(p.get('compatible_model') or '').strip(),
                quantity=safe_int(p.get('quantity'), 0),
                selling_price=safe_float(p.get('selling_price'), 0),
                low_stock_limit=safe_int(p.get('low_stock_limit'), 0),
            )
            if part.name:
                db.session.add(part)

        for item in payload.get('repairs', []):
            cust = customer_map.get(((item.get('customer_name') or '').strip(), (item.get('customer_phone') or '').strip()))
            if not cust:
                continue
            repair = RepairJob(
                company_id=company.id,
                repair_code=(item.get('repair_code') or generate_repair_code(company.id)).strip(),
                invoice_number=(item.get('invoice_number') or generate_invoice_number(company.id)).strip(),
                customer_id=cust.id,
                device_brand=(item.get('device_brand') or '').strip(),
                device_model=(item.get('device_model') or '').strip(),
                serial_number=(item.get('serial_number') or '').strip(),
                imei=(item.get('imei') or '').strip(),
                issue_description=(item.get('issue_description') or '').strip(),
                status=(item.get('status') or 'Received').strip(),
                estimated_cost=safe_float(item.get('estimated_cost'), 0),
                deposit_amount=safe_float(item.get('deposit_amount'), 0),
                due_date=parse_date_field(item.get('due_date')),
                customer_note=(item.get('customer_note') or '').strip(),
                items_json=json.dumps(item.get('items') or [{
                    'brand': (item.get('device_brand') or '').strip(),
                    'model': (item.get('device_model') or '').strip(),
                    'serial': (item.get('serial_number') or '').strip(),
                    'imei': (item.get('imei') or '').strip(),
                    'issue': (item.get('issue_description') or '').strip(),
                    'amount': safe_float(item.get('estimated_cost'), 0),
                }]),
            )
            db.session.add(repair)
            db.session.flush()
            for p in item.get('payments', []):
                db.session.add(Payment(company_id=company.id, repair_job_id=repair.id, amount=safe_float(p.get('amount'), 0), method=(p.get('method') or 'Cash').strip() or 'Cash'))
            for n in item.get('internal_notes', []):
                db.session.add(InternalNote(company_id=company.id, repair_job_id=repair.id, user_id=current_user.id, note=(n.get('note') or '').strip()))
        for item in payload.get('gsm_tickets', []):
            cust = customer_map.get(((item.get('customer_name') or '').strip(), (item.get('customer_phone') or '').strip()))
            if not cust:
                continue
            bt = BuyingTicket(company_id=company.id, ticket_code=(item.get('ticket_code') or generate_ticket_code(company.id)).strip(), invoice_number=(item.get('invoice_number') or generate_buying_invoice_number(company.id)).strip(), customer_id=cust.id, device_brand=(item.get('device_brand') or '').strip(), device_model=(item.get('device_model') or '').strip(), storage_gb=(item.get('storage_gb') or '').strip(), serial_number=(item.get('serial_number') or '').strip(), imei=(item.get('imei') or '').strip(), battery_percentage=(item.get('battery_percentage') or '').strip(), issue_description=(item.get('note') or '').strip(), status=(item.get('status') or 'Buying').strip(), estimated_cost=safe_float(item.get('estimated_cost'),0), deposit_amount=0, due_date=parse_date_field(item.get('due_date')), customer_note='')
            db.session.add(bt); db.session.flush()
            for p in item.get('payments', []):
                db.session.add(BuyingPayment(company_id=company.id, buying_ticket_id=bt.id, amount=safe_float(p.get('amount'), 0), method=(p.get('method') or 'Cash').strip() or 'Cash'))
        db.session.commit()
        return True, 'Full backup imported successfully.'

    @app.before_request
    def set_request_language():
        g.lang = current_language()

    @app.context_processor
    def inject_globals():
        company = None if not current_user.is_authenticated or current_user.is_super_admin else current_user.company
        settings = AppSetting.query.first()
        if settings and (not getattr(settings, 'menu_buying', None) or str(settings.menu_buying).strip() == '' or settings.menu_buying == 'Buying'):
            settings.menu_buying = 'GSM'
        return {
            'role_labels': ROLE_LABELS,
            'company_role_labels': COMPANY_ROLE_LABELS,
            'statuses': get_statuses(),
            'payment_methods': get_payment_methods(),
            'active_company': company,
            't': t,
            'current_lang': getattr(g, 'lang', 'en'),
            'status_label': status_label,
            'today_iso': date.today().isoformat(),
            'app_settings': settings,
            'platform_name': (settings.platform_name if settings and settings.platform_name else 'Repair Control'),
            'is_impersonating': bool(session.get('impersonator_id')),
            'label_or_t': lambda current, default, key: (t(key) if (not current or current == default) else current),
            'ticket_items': ticket_items,
            'format_gsm_description': format_gsm_description,
            'repair_items': repair_items,
            'format_repair_description': format_repair_description,
        }

    @app.route('/')
    def index():
        if current_user.is_authenticated:
            return redirect(url_for('super_admin_dashboard' if current_user.is_super_admin else 'dashboard'))
        return redirect(url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            email = request.form['email'].strip().lower()
            password = request.form['password']
            user = User.query.filter_by(email=email).first()
            if user and check_password_hash(user.password_hash, password):
                allowed, reason = is_account_allowed(user)
                if not allowed:
                    flash(reason, 'danger')
                else:
                    login_user(user)
                    return redirect(url_for('super_admin_dashboard' if user.is_super_admin else 'dashboard'))
            else:
                flash('Invalid email or password', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    # ---------- Super Admin ----------
    @app.route('/super-admin')
    @roles_required('Super Admin')
    def super_admin_dashboard():
        companies = Company.query.order_by(Company.created_at.desc()).limit(8).all()
        active_companies = Company.query.filter_by(is_active=True).count()
        expired_companies = Company.query.filter(Company.unlimited_access.is_(False), Company.active_until.isnot(None), Company.active_until < date.today()).count()
        attention_companies = Company.query.filter(or_(Company.is_active.is_(False), (Company.unlimited_access.is_(False) & Company.active_until.isnot(None) & (Company.active_until <= date.today() + timedelta(days=7))))).count()
        company_stats = {c.id: {'users': User.query.filter_by(company_id=c.id).count(), 'repairs': RepairJob.query.filter_by(company_id=c.id, is_archived=False).count()} for c in companies}
        return render_template(
            'super_admin_dashboard.html',
            companies=companies,
            company_stats=company_stats,
            companies_count=Company.query.count(),
            active_companies=active_companies,
            expired_companies=expired_companies,
            attention_companies=attention_companies,
            users_count=User.query.count(),
            repairs_count=RepairJob.query.count(),
            status_count=StatusOption.query.count(),
            payment_method_count=PaymentMethod.query.count(),
            recent_audit=AuditLog.query.order_by(AuditLog.created_at.desc()).limit(10).all(),
        )

    @app.route('/super-admin/companies', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_companies():
        if request.method == 'POST':
            name = request.form['name'].strip()
            phone = normalize_phone(request.form.get('phone', '').strip())
            email = request.form.get('email', '').strip().lower()
            address = request.form.get('address', '').strip()
            admin_name = request.form['admin_name'].strip()
            admin_email = request.form['admin_email'].strip().lower()
            admin_password = request.form['admin_password']
            unlimited_access = bool(request.form.get('unlimited_access'))
            active_until = parse_date_field(request.form.get('active_until'))
            plan_type = request.form.get('plan_type', 'trial')
            if not all([name, admin_name, admin_email, admin_password]):
                flash('Please fill in all required company fields.', 'danger')
                return redirect(url_for('super_admin_companies'))
            if not unlimited_access and not active_until:
                flash('Please choose an active until date or select unlimited access.', 'danger')
                return redirect(url_for('super_admin_companies'))
            if Company.query.filter_by(name=name).first():
                flash('A company with that name already exists.', 'danger')
                return redirect(url_for('super_admin_companies'))
            if User.query.filter_by(email=admin_email).first():
                flash('That admin email already exists.', 'danger')
                return redirect(url_for('super_admin_companies'))
            settings = AppSetting.query.first()
            company_default_language = settings.platform_language if settings and settings.platform_language else 'en'
            company = Company(name=name, phone=phone, email=email, address=address, plan_type=plan_type, unlimited_access=unlimited_access, active_until=None if unlimited_access else active_until, default_language=company_default_language)
            db.session.add(company)
            db.session.flush()
            admin_user = User(company_id=company.id, name=admin_name, email=admin_email, password_hash=generate_password_hash(admin_password), role='Company Admin', is_active=True, unlimited_access=True, active_until=None)
            db.session.add(admin_user)
            log_audit('company_created', 'Company', company.id, f'{company.name} created with plan {company.plan_type}.')
            log_audit('user_created', 'User', admin_user.id, f'Initial company admin {admin_user.email} created for {company.name}.')
            db.session.commit()
            flash('Company and company admin created.', 'success')
            return redirect(url_for('super_admin_companies'))
        q = request.args.get('q', '').strip()
        query = Company.query
        if q:
            like = f'%{q}%'
            query = query.filter(or_(Company.name.ilike(like), Company.email.ilike(like), Company.phone.ilike(like), Company.address.ilike(like)))
        companies = query.order_by(Company.created_at.desc()).all()
        company_stats = {c.id: {'users': User.query.filter_by(company_id=c.id).count(), 'repairs': RepairJob.query.filter_by(company_id=c.id, is_archived=False).count()} for c in companies}
        return render_template('super_admin_companies.html', companies=companies, company_stats=company_stats, q=q)

    @app.route('/super-admin/companies/<int:company_id>/toggle', methods=['POST'])
    @roles_required('Super Admin')
    def toggle_company(company_id):
        company = get_company_or_404(company_id)
        company.is_active = not company.is_active
        log_audit('company_toggled', 'Company', company.id, f'Company {company.name} set to {'active' if company.is_active else 'disabled'}.')
        db.session.commit()
        flash(f'Company {"enabled" if company.is_active else "disabled"}.', 'success')
        return redirect(url_for('super_admin_companies'))

    @app.route('/super-admin/companies/<int:company_id>/access', methods=['POST'])
    @roles_required('Super Admin')
    def super_admin_company_access(company_id):
        company = get_company_or_404(company_id)
        company.plan_type = request.form.get('plan_type', company.plan_type)
        company.unlimited_access = bool(request.form.get('unlimited_access'))
        company.active_until = None if company.unlimited_access else parse_date_field(request.form.get('active_until'))
        log_audit('company_access_updated', 'Company', company.id, f'Plan={company.plan_type}, unlimited={company.unlimited_access}, active_until={company.active_until}.')
        db.session.commit()
        flash('Company access updated.', 'success')
        return redirect(request.referrer or url_for('super_admin_companies'))

    @app.route('/super-admin/companies/<int:company_id>/impersonate', methods=['POST'])
    @roles_required('Super Admin')
    def super_admin_impersonate(company_id):
        company = get_company_or_404(company_id)
        target = User.query.filter_by(company_id=company.id, role='Company Admin', is_active=True).order_by(User.id.asc()).first()
        if not target:
            target = User.query.filter_by(company_id=company.id, is_active=True).order_by(User.id.asc()).first()
        if not target:
            flash('No active user exists for that company.', 'danger')
            return redirect(url_for('super_admin_companies'))
        session['impersonator_id'] = current_user.id
        log_audit('company_impersonated', 'Company', company.id, f'Super admin impersonated {company.name}.')
        login_user(target)
        flash(f'You are now viewing the app as {company.name}.', 'success')
        return redirect(url_for('dashboard'))

    @app.route('/super-admin/stop-impersonation', methods=['POST'])
    @login_required
    def stop_impersonation():
        impersonator_id = session.pop('impersonator_id', None)
        if not impersonator_id:
            return redirect(url_for('dashboard' if not current_user.is_super_admin else 'super_admin_dashboard'))
        admin = User.query.get(impersonator_id)
        if not admin or not admin.is_super_admin:
            flash('Could not restore your super admin session.', 'danger')
            return redirect(url_for('login'))
        login_user(admin)
        flash('Returned to Super Admin.', 'success')
        return redirect(url_for('super_admin_dashboard'))

    @app.route('/super-admin/profile')
    @roles_required('Super Admin')
    def super_admin_profile():
        return redirect(url_for('profile'))

    @app.route('/profile', methods=['GET', 'POST'])
    @login_required
    def profile():
        user = current_user
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            email = request.form.get('email', '').strip().lower()
            current_password = request.form.get('current_password', '')
            password = request.form.get('password', '')
            confirm_password = request.form.get('confirm_password', '')
            if not name or not email:
                flash('Name and email are required.', 'danger')
                return redirect(url_for('profile'))
            existing = User.query.filter(User.email == email, User.id != user.id).first()
            if existing:
                flash('That email is already in use.', 'danger')
                return redirect(url_for('profile'))
            if password or confirm_password:
                if not current_password or not check_password_hash(user.password_hash, current_password):
                    flash('Current password is incorrect.', 'danger')
                    return redirect(url_for('profile'))
                if password != confirm_password:
                    flash('Passwords do not match.', 'danger')
                    return redirect(url_for('profile'))
                user.password_hash = generate_password_hash(password)
            user.name = name
            user.email = email
            db.session.commit()
            flash('Profile updated.', 'success')
            return redirect(url_for('profile'))
        return render_template('profile.html', user=current_user)

    @app.route('/super-admin/master-settings', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_master_settings():
        settings = AppSetting.query.first()
        if request.method == 'POST':
            settings.platform_name = request.form.get('platform_name', '').strip() or 'Repair Control'
            settings.platform_language = request.form.get('platform_language', 'en').strip() or 'en'
            settings.theme_primary = request.form.get('theme_primary', '#38bdf8')
            settings.theme_accent = request.form.get('theme_accent', '#0f172a')
            settings.support_enabled = bool(request.form.get('support_enabled'))
            settings.support_whatsapp = normalize_phone(request.form.get('support_whatsapp', '').strip())
            settings.support_message = request.form.get('support_message', '').strip()
            settings.support_embed_code = request.form.get('support_embed_code', '').strip()
            settings.support_custom_css = request.form.get('support_custom_css', '').strip()
            settings.invoice_title = request.form.get('invoice_title', '').strip() or 'Repair Invoice'
            settings.invoice_line_label = request.form.get('invoice_line_label', '').strip() or 'Problem Description'
            settings.qr_caption = request.form.get('qr_caption', '').strip() or 'Scan for status'
            settings.show_public_status_url = bool(request.form.get('show_public_status_url'))
            settings.menu_dashboard = request.form.get('menu_dashboard', '').strip() or 'Dashboard'
            settings.menu_customers = request.form.get('menu_customers', '').strip() or 'Customers'
            settings.menu_parts = request.form.get('menu_parts', '').strip() or 'Parts'
            settings.menu_repairs = request.form.get('menu_repairs', '').strip() or 'Repairs'
            settings.menu_buying = request.form.get('menu_buying', '').strip() or 'GSM'
            settings.menu_reports = request.form.get('menu_reports', '').strip() or 'Reports'
            settings.menu_audit_log = request.form.get('menu_audit_log', '').strip() or 'Audit Log'
            settings.menu_users = request.form.get('menu_users', '').strip() or 'Users'
            settings.menu_settings = request.form.get('menu_settings', '').strip() or 'Settings'
            settings.menu_data_manager = request.form.get('menu_data_manager', '').strip() or 'Data Manager'
            settings.menu_profile = request.form.get('menu_profile', '').strip() or 'Profile'
            settings.menu_logout = request.form.get('menu_logout', '').strip() or 'Logout'
            settings.menu_companies = request.form.get('menu_companies', '').strip() or 'Companies'
            settings.menu_status_manager = request.form.get('menu_status_manager', '').strip() or 'Status Manager'
            settings.menu_payment_methods = request.form.get('menu_payment_methods', '').strip() or 'Payment Methods'
            settings.menu_master_settings = request.form.get('menu_master_settings', '').strip() or 'Master Settings'
            settings.label_repairs_count = request.form.get('label_repairs_count', '').strip() or 'Repairs'
            settings.label_customers_count = request.form.get('label_customers_count', '').strip() or 'Customers'
            settings.label_parts_count = request.form.get('label_parts_count', '').strip() or 'Parts'
            settings.label_users_count = request.form.get('label_users_count', '').strip() or 'Users'
            settings.label_low_stock = request.form.get('label_low_stock', '').strip() or 'Low Stock'
            settings.label_overdue_repairs = request.form.get('label_overdue_repairs', '').strip() or 'Overdue Repairs'
            settings.label_todays_repairs = request.form.get('label_todays_repairs', '').strip() or "Today's repairs"
            settings.label_recent_payments = request.form.get('label_recent_payments', '').strip() or 'Recent payments'
            settings.label_recent_activity = request.form.get('label_recent_activity', '').strip() or 'Recent activity'
            settings.label_recent_repairs = request.form.get('label_recent_repairs', '').strip() or 'Recent Repairs'
            settings.label_search_placeholder = request.form.get('label_search_placeholder', '').strip() or 'Search by phone / repair ID / serial / IMEI / customer'
            settings.label_repair_code = request.form.get('label_repair_code', '').strip() or 'Repair Code'
            settings.label_customer = request.form.get('label_customer', '').strip() or 'Customer'
            settings.label_device = request.form.get('label_device', '').strip() or 'Device'
            settings.label_status = request.form.get('label_status', '').strip() or 'Status'
            settings.label_action = request.form.get('label_action', '').strip() or 'Action'
            settings.label_open = request.form.get('label_open', '').strip() or 'Open'
            settings.gsm_bought_status_label = request.form.get('gsm_bought_status_label', '').strip() or 'BOUGHT BY SMP'
            settings.gsm_sold_status_label = request.form.get('gsm_sold_status_label', '').strip() or 'SOLD BY SMP'
            log_audit('master_settings_updated', 'AppSetting', settings.id, f"Global theme/support/invoice settings updated. Language set to {settings.platform_language}.")
            db.session.commit()
            flash('Master settings updated.', 'success')
            return redirect(url_for('super_admin_master_settings'))
        return render_template('super_admin_master_settings.html', settings=settings)

    @app.route('/super-admin/statuses', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_statuses():
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('Status name is required.', 'danger')
                return redirect(url_for('super_admin_statuses'))
            if StatusOption.query.filter_by(name=name).first():
                flash('That status already exists.', 'danger')
                return redirect(url_for('super_admin_statuses'))
            next_order = (db.session.query(func.max(StatusOption.sort_order)).scalar() or 0) + 1
            db.session.add(StatusOption(name=name, sort_order=next_order, is_active=True))
            db.session.commit()
            flash('Status created.', 'success')
            return redirect(url_for('super_admin_statuses'))
        statuses = StatusOption.query.order_by(StatusOption.sort_order.asc(), StatusOption.name.asc()).all()
        return render_template('super_admin_statuses.html', statuses=statuses)

    @app.route('/super-admin/statuses/<int:status_id>/toggle', methods=['POST'])
    @roles_required('Super Admin')
    def toggle_status(status_id):
        status = StatusOption.query.get_or_404(status_id)
        status.is_active = not status.is_active
        log_audit('status_toggled', 'StatusOption', status.id, f'Status {status.name} set to {status.is_active}.')
        db.session.commit()
        flash('Status updated.', 'success')
        return redirect(url_for('super_admin_statuses'))

    @app.route('/super-admin/payment-methods', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_payment_methods():
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('Payment method name is required.', 'danger')
                return redirect(url_for('super_admin_payment_methods'))
            if PaymentMethod.query.filter_by(name=name).first():
                flash('That payment method already exists.', 'danger')
                return redirect(url_for('super_admin_payment_methods'))
            next_order = (db.session.query(func.max(PaymentMethod.sort_order)).scalar() or 0) + 1
            db.session.add(PaymentMethod(name=name, sort_order=next_order, is_active=True))
            db.session.commit()
            flash('Payment method created.', 'success')
            return redirect(url_for('super_admin_payment_methods'))
        methods = PaymentMethod.query.order_by(PaymentMethod.sort_order.asc(), PaymentMethod.name.asc()).all()
        return render_template('super_admin_payment_methods.html', methods=methods)

    @app.route('/super-admin/payment-methods/<int:method_id>/toggle', methods=['POST'])
    @roles_required('Super Admin')
    def toggle_payment_method(method_id):
        method = PaymentMethod.query.get_or_404(method_id)
        method.is_active = not method.is_active
        log_audit('payment_method_toggled', 'PaymentMethod', method.id, f'Payment method {method.name} set to {method.is_active}.')
        db.session.commit()
        flash('Payment method updated.', 'success')
        return redirect(url_for('super_admin_payment_methods'))

    @app.route('/super-admin/companies/<int:company_id>/settings', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_company_settings(company_id):
        company = get_company_or_404(company_id)
        if request.method == 'POST':
            company.name = request.form['name'].strip()
            company.phone = normalize_phone(request.form.get('phone', '').strip())
            company.email = request.form.get('email', '').strip()
            company.address = request.form.get('address', '').strip()
            company.vat_number = request.form.get('vat_number', '').strip()
            company.invoice_terms_en = request.form.get('invoice_terms_en', '').strip()
            company.invoice_terms_fr = request.form.get('invoice_terms_fr', '').strip()
            company.invoice_terms_nl = request.form.get('invoice_terms_nl', '').strip()
            company.storage_notice_en = request.form.get('storage_notice_en', '').strip()
            company.storage_notice_fr = request.form.get('storage_notice_fr', '').strip()
            company.storage_notice_nl = request.form.get('storage_notice_nl', '').strip()
            company.gsm_invoice_title = request.form.get('gsm_invoice_title', '').strip() or 'Ticket'
            company.gsm_terms_en = request.form.get('gsm_terms_en', '').strip()
            company.gsm_terms_fr = request.form.get('gsm_terms_fr', '').strip()
            company.gsm_terms_nl = request.form.get('gsm_terms_nl', '').strip()
            company.gsm_notice_en = request.form.get('gsm_notice_en', '').strip()
            company.gsm_notice_fr = request.form.get('gsm_notice_fr', '').strip()
            company.gsm_notice_nl = request.form.get('gsm_notice_nl', '').strip()
            company.invoice_terms = company.invoice_terms_en
            company.storage_notice = company.storage_notice_en
            company.default_language = request.form.get('default_language', 'en')
            if current_user.is_super_admin:
                company.plan_type = request.form.get('plan_type', company.plan_type)
                company.unlimited_access = bool(request.form.get('unlimited_access'))
                company.active_until = None if company.unlimited_access else parse_date_field(request.form.get('active_until'))
                company.is_active = bool(request.form.get('is_active'))
            company.print_width_in = float(request.form.get('print_width_in') or company.print_width_in or 3.82)
            company.print_height_in = float(request.form.get('print_height_in') or company.print_height_in or 5.80)
            company.print_padding_x_in = float(request.form.get('print_padding_x_in') or company.print_padding_x_in or 0.12)
            company.print_padding_y_in = float(request.form.get('print_padding_y_in') or company.print_padding_y_in or 0.12)
            company.print_header_font_px = int(request.form.get('print_header_font_px') or company.print_header_font_px or 10)
            company.print_body_font_px = int(request.form.get('print_body_font_px') or company.print_body_font_px or 11)
            company.print_notice_font_px = int(request.form.get('print_notice_font_px') or company.print_notice_font_px or 9)
            company.print_content_shift_in = float(request.form.get('print_content_shift_in') or company.print_content_shift_in or 0.18)
            company.print_80mm_width_mm = float(request.form.get('print_80mm_width_mm') or company.print_80mm_width_mm or 80)
            company.print_80mm_padding_mm = float(request.form.get('print_80mm_padding_mm') or company.print_80mm_padding_mm or 5)
            company.print_80mm_header_font_px = int(request.form.get('print_80mm_header_font_px') or company.print_80mm_header_font_px or 11)
            company.print_80mm_body_font_px = int(request.form.get('print_80mm_body_font_px') or company.print_80mm_body_font_px or 10)
            company.print_80mm_notice_font_px = int(request.form.get('print_80mm_notice_font_px') or company.print_80mm_notice_font_px or 9)
            company.print_80mm_content_shift_mm = float(request.form.get('print_80mm_content_shift_mm') or company.print_80mm_content_shift_mm or 0)
            logo_file = request.files.get('logo')
            logo_path, logo_error = save_logo(logo_file)
            if logo_error:
                flash(logo_error, 'danger')
                return redirect(request.url)
            if logo_path:
                company.logo_path = logo_path
            log_audit('company_settings_updated', 'Company', company.id, f'Settings updated for {company.name}.')
            db.session.commit()
            flash('Company settings updated.', 'success')
            return redirect(url_for('super_admin_company_settings', company_id=company.id))
        return render_template('company_settings.html', company=company, super_admin_mode=True)

    @app.route('/super-admin/companies/<int:company_id>/users', methods=['GET', 'POST'])
    @roles_required('Super Admin')
    def super_admin_company_users(company_id):
        company = get_company_or_404(company_id)
        if request.method == 'POST':
            name = request.form['name'].strip()
            email = request.form['email'].strip().lower()
            password = request.form['password']
            role = request.form.get('role', 'Staff')
            unlimited_access = bool(request.form.get('unlimited_access'))
            active_until = parse_date_field(request.form.get('active_until'))
            if role not in COMPANY_ROLE_LABELS:
                role = 'Staff'
            if not unlimited_access and not active_until:
                flash('Please choose an active until date or select unlimited access.', 'danger')
                return redirect(url_for('super_admin_company_users', company_id=company.id))
            if User.query.filter_by(email=email).first():
                flash('That email already exists.', 'danger')
                return redirect(url_for('super_admin_company_users', company_id=company.id))
            user = User(company_id=company.id, name=name, email=email, password_hash=generate_password_hash(password), role=role, is_active=True, unlimited_access=unlimited_access, active_until=None if unlimited_access else active_until)
            db.session.add(user)
            log_audit('user_created', 'User', None, f'User {email} created for {company.name} as {role}.')
            db.session.commit()
            flash('Company user created.', 'success')
            return redirect(url_for('super_admin_company_users', company_id=company.id))
        users = User.query.filter_by(company_id=company.id).order_by(User.created_at.desc()).all()
        return render_template('company_users.html', company=company, users=users, super_admin_mode=True)

    @app.route('/super-admin/users/<int:user_id>/toggle', methods=['POST'])
    @roles_required('Super Admin')
    def super_admin_toggle_user(user_id):
        user = User.query.get_or_404(user_id)
        if user.is_super_admin:
            flash('Super Admin cannot be disabled here.', 'danger')
            return redirect(url_for('super_admin_dashboard'))
        user.is_active = not user.is_active
        log_audit('user_toggled', 'User', user.id, f'User {user.email} set to {user.is_active}.')
        db.session.commit()
        flash(f'User {"enabled" if user.is_active else "disabled"}.', 'success')
        return redirect(url_for('super_admin_company_users', company_id=user.company_id))

    # ---------- Company Area ----------
    @app.route('/dashboard')
    @company_required
    def dashboard():
        q = request.args.get('q', '').strip()
        jobs_query = company_query(RepairJob).join(Customer)
        if q:
            jobs_query = jobs_query.filter(
                or_(
                    RepairJob.repair_code.ilike(f'%{q}%'),
                    RepairJob.invoice_number.ilike(f'%{q}%'),
                    RepairJob.device_brand.ilike(f'%{q}%'),
                    RepairJob.device_model.ilike(f'%{q}%'),
                    RepairJob.serial_number.ilike(f'%{q}%'),
                    RepairJob.imei.ilike(f'%{q}%'),
                    Customer.name.ilike(f'%{q}%'),
                    Customer.phone.ilike(f'%{q}%'),
                )
            )
        jobs = jobs_query.order_by(RepairJob.created_at.desc()).limit(25).all()
        company_id = current_user.company_id
        status_counts = dict(db.session.query(RepairJob.status, func.count(RepairJob.id)).filter_by(company_id=company_id, is_archived=False).group_by(RepairJob.status).all())
        overdue_repairs = RepairJob.query.filter(RepairJob.company_id == company_id, RepairJob.is_archived.is_(False), RepairJob.due_date.isnot(None), RepairJob.due_date < date.today(), ~RepairJob.status.in_(['Delivered', 'Cancelled'])).count()
        low_stock = company_query(Part).filter(Part.quantity <= Part.low_stock_limit).count()
        todays_repairs = RepairJob.query.filter(RepairJob.company_id == company_id, RepairJob.is_archived.is_(False), func.date(RepairJob.created_at) == date.today()).count()
        total_gsm = company_query(BuyingTicket).count()
        todays_gsm = BuyingTicket.query.filter(BuyingTicket.company_id == company_id, func.date(BuyingTicket.created_at) == date.today()).count()
        recent_gsm = company_query(BuyingTicket).order_by(BuyingTicket.created_at.desc()).limit(5).all()
        recent_payments = Payment.query.filter_by(company_id=company_id).order_by(Payment.created_at.desc()).limit(5).all()
        recent_updates = RepairUpdate.query.filter_by(company_id=company_id).order_by(RepairUpdate.created_at.desc()).limit(5).all()
        recent_notes = InternalNote.query.filter_by(company_id=company_id).order_by(InternalNote.created_at.desc()).limit(5).all()
        recent_activity = sorted(
            ([{'when': x.created_at, 'kind':'payment', 'title':f'Payment € {x.amount:.2f}', 'detail':x.method, 'repair_id':x.repair_job_id} for x in recent_payments]
            + [{'when': x.created_at, 'kind':'status', 'title':f'{x.old_status or '-'} → {x.new_status}', 'detail': x.note or '', 'repair_id':x.repair_job_id} for x in recent_updates]
            + [{'when': x.created_at, 'kind':'note', 'title':'Internal note', 'detail':x.note, 'repair_id':x.repair_job_id} for x in recent_notes]),
            key=lambda a: a['when'], reverse=True
        )[:8]
        return render_template(
            'dashboard.html',
            jobs=jobs,
            q=q,
            total_jobs=company_query(RepairJob).count(),
            total_customers=company_query(Customer).count(),
            total_parts=company_query(Part).count(),
            total_users=User.query.filter_by(company_id=current_user.company_id).count(),
            low_stock=low_stock,
            overdue_repairs=overdue_repairs,
            todays_repairs=todays_repairs,
            total_gsm=total_gsm,
            todays_gsm=todays_gsm,
            recent_gsm=recent_gsm,
            status_counts=status_counts,
            recent_payments=recent_payments,
            recent_activity=recent_activity,
        )

    @app.route('/settings', methods=['GET', 'POST'])
    @roles_required('Company Admin')
    @company_required
    def company_settings():
        company = current_user.company
        if request.method == 'POST':
            company.name = request.form['name'].strip()
            company.phone = normalize_phone(request.form.get('phone', '').strip())
            company.email = request.form.get('email', '').strip()
            company.address = request.form.get('address', '').strip()
            company.vat_number = request.form.get('vat_number', '').strip()
            company.invoice_terms_en = request.form.get('invoice_terms_en', '').strip()
            company.invoice_terms_fr = request.form.get('invoice_terms_fr', '').strip()
            company.invoice_terms_nl = request.form.get('invoice_terms_nl', '').strip()
            company.storage_notice_en = request.form.get('storage_notice_en', '').strip()
            company.storage_notice_fr = request.form.get('storage_notice_fr', '').strip()
            company.storage_notice_nl = request.form.get('storage_notice_nl', '').strip()
            company.gsm_invoice_title = request.form.get('gsm_invoice_title', '').strip() or 'Ticket'
            company.gsm_terms_en = request.form.get('gsm_terms_en', '').strip()
            company.gsm_terms_fr = request.form.get('gsm_terms_fr', '').strip()
            company.gsm_terms_nl = request.form.get('gsm_terms_nl', '').strip()
            company.gsm_notice_en = request.form.get('gsm_notice_en', '').strip()
            company.gsm_notice_fr = request.form.get('gsm_notice_fr', '').strip()
            company.gsm_notice_nl = request.form.get('gsm_notice_nl', '').strip()
            company.invoice_terms = company.invoice_terms_en
            company.storage_notice = company.storage_notice_en
            company.default_language = request.form.get('default_language', 'en')
            if current_user.is_super_admin:
                company.plan_type = request.form.get('plan_type', company.plan_type)
                company.unlimited_access = bool(request.form.get('unlimited_access'))
                company.active_until = None if company.unlimited_access else parse_date_field(request.form.get('active_until'))
                company.is_active = bool(request.form.get('is_active'))
            company.print_width_in = float(request.form.get('print_width_in') or company.print_width_in or 3.82)
            company.print_height_in = float(request.form.get('print_height_in') or company.print_height_in or 5.80)
            company.print_padding_x_in = float(request.form.get('print_padding_x_in') or company.print_padding_x_in or 0.12)
            company.print_padding_y_in = float(request.form.get('print_padding_y_in') or company.print_padding_y_in or 0.12)
            company.print_header_font_px = int(request.form.get('print_header_font_px') or company.print_header_font_px or 10)
            company.print_body_font_px = int(request.form.get('print_body_font_px') or company.print_body_font_px or 11)
            company.print_notice_font_px = int(request.form.get('print_notice_font_px') or company.print_notice_font_px or 9)
            company.print_content_shift_in = float(request.form.get('print_content_shift_in') or company.print_content_shift_in or 0.18)
            company.print_80mm_width_mm = float(request.form.get('print_80mm_width_mm') or company.print_80mm_width_mm or 80)
            company.print_80mm_padding_mm = float(request.form.get('print_80mm_padding_mm') or company.print_80mm_padding_mm or 5)
            company.print_80mm_header_font_px = int(request.form.get('print_80mm_header_font_px') or company.print_80mm_header_font_px or 11)
            company.print_80mm_body_font_px = int(request.form.get('print_80mm_body_font_px') or company.print_80mm_body_font_px or 10)
            company.print_80mm_notice_font_px = int(request.form.get('print_80mm_notice_font_px') or company.print_80mm_notice_font_px or 9)
            company.print_80mm_content_shift_mm = float(request.form.get('print_80mm_content_shift_mm') or company.print_80mm_content_shift_mm or 0)
            logo_file = request.files.get('logo')
            logo_path, logo_error = save_logo(logo_file)
            if logo_error:
                flash(logo_error, 'danger')
                return redirect(request.url)
            if logo_path:
                company.logo_path = logo_path
            log_audit('company_settings_updated', 'Company', company.id, f'Company admin updated settings for {company.name}.')
            db.session.commit()
            flash('Company settings updated.', 'success')
            return redirect(url_for('company_settings'))
        return render_template('company_settings.html', company=company, super_admin_mode=False)

    @app.route('/data-manager', methods=['GET'])
    @company_required
    @roles_required('Company Admin')
    def data_manager():
        return render_template('data_manager.html')

    @app.route('/data/export/full-json')
    @company_required
    @roles_required('Company Admin')
    def export_full_json():
        payload = export_company_json(current_user.company)
        filename = f'{current_user.company.name.lower().replace(" ","_")}_backup.json'
        return Response(json.dumps(payload, ensure_ascii=False, indent=2), mimetype='application/json', headers={'Content-Disposition': f'attachment; filename={filename}'})

    @app.route('/data/export/full-xlsx')
    @company_required
    @roles_required('Company Admin')
    def export_full_xlsx():
        buf = export_company_xlsx(current_user.company)
        filename = f'{current_user.company.name.lower().replace(" ","_")}_backup.xlsx'
        return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}'})

    @app.route('/data/export/<kind>')
    @company_required
    @roles_required('Company Admin')
    def export_data(kind):
        export_format = request.args.get('format', 'csv').lower()
        rows = []
        headers = []
        filename = f'{kind}.csv'
        if kind == 'customers':
            headers = ['name', 'phone', 'email']
            rows = [[row.name, row.phone, row.email or ''] for row in company_query(Customer).order_by(Customer.created_at.asc()).all()]
            filename = 'customers'
        elif kind == 'parts':
            headers = ['name', 'compatible_model', 'quantity', 'selling_price', 'low_stock_limit']
            rows = [[row.name, row.compatible_model or '', row.quantity, row.selling_price, row.low_stock_limit] for row in company_query(Part).order_by(Part.created_at.asc()).all()]
            filename = 'parts'
        elif kind == 'repairs':
            headers = ['repair_code', 'invoice_number', 'customer_name', 'customer_phone', 'device_brand', 'device_model', 'serial_number', 'imei', 'status', 'estimated_cost', 'deposit_amount', 'due_date']
            rows = [[row.repair_code, row.invoice_number, row.customer.name, row.customer.phone, row.device_brand, row.device_model, row.serial_number or '', row.imei or '', row.status, row.estimated_cost, row.deposit_amount, row.due_date or ''] for row in company_query(RepairJob).join(Customer).order_by(RepairJob.created_at.asc()).all()]
            filename = 'repairs'
        else:
            abort(404)
        if export_format == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = kind
            ws.append(headers)
            for row in rows:
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'})
        sio = io.StringIO()
        writer = csv.writer(sio)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return Response(sio.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename}.csv'})

    @app.route('/data/import/full-json', methods=['POST'])
    @company_required
    @roles_required('Company Admin')
    def import_full_json():
        ok, message = import_company_json(request.files.get('file'), current_user.company)
        flash(message, 'success' if ok else 'danger')
        return redirect(url_for('data_manager'))

    @app.route('/data/import/full-xlsx', methods=['POST'])
    @company_required
    @roles_required('Company Admin')
    def import_full_xlsx():
        ok, message = import_company_xlsx(request.files.get('file'), current_user.company)
        flash(message, 'success' if ok else 'danger')
        return redirect(url_for('data_manager'))

    @app.route('/export/backup')
    @company_required
    @roles_required('Company Admin')
    def export_backup():
        return redirect(url_for('data_manager'))

    @app.route('/users', methods=['GET', 'POST'])
    @roles_required('Company Admin')
    @company_required
    def users():
        if request.method == 'POST':
            name = request.form['name'].strip()
            email = request.form['email'].strip().lower()
            password = request.form['password']
            role = request.form.get('role', 'Staff')
            if role not in COMPANY_ROLE_LABELS:
                role = 'Staff'
            if User.query.filter_by(email=email).first():
                flash('That email already exists.', 'danger')
                return redirect(url_for('users'))
            user = User(company_id=current_user.company_id, name=name, email=email, password_hash=generate_password_hash(password), role=role, is_active=True, unlimited_access=True, active_until=None)
            db.session.add(user)
            log_audit('user_created', 'User', None, f'Company user {email} created as {role}.')
            db.session.commit()
            flash('User account created.', 'success')
            return redirect(url_for('users'))
        users = User.query.filter_by(company_id=current_user.company_id).order_by(User.created_at.desc()).all()
        return render_template('company_users.html', company=current_user.company, users=users, super_admin_mode=False)

    @app.route('/users/<int:user_id>/toggle', methods=['POST'])
    @roles_required('Company Admin')
    @company_required
    def toggle_user(user_id):
        user = company_object_or_404(User, user_id)
        if user.role == 'Company Admin' and user.id == current_user.id:
            flash('You cannot disable your own company admin account.', 'danger')
            return redirect(url_for('users'))
        user.is_active = not user.is_active
        log_audit('user_toggled', 'User', user.id, f'Company user {user.email} set to {user.is_active}.')
        db.session.commit()
        flash(f'User {"enabled" if user.is_active else "disabled"}.', 'success')
        return redirect(url_for('users'))

    @app.route('/customers', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff')
    def customers():
        if request.method == 'POST':
            name = request.form['name'].strip()
            phone = normalize_phone(request.form['phone'].strip())
            email = request.form.get('email', '').strip()
            if not name or not phone:
                flash('Customer name and phone are required.', 'danger')
                return redirect(url_for('customers'))
            if Customer.query.filter_by(company_id=current_user.company_id, name=name, phone=phone).first():
                flash('That customer already exists.', 'danger')
                return redirect(url_for('customers'))
            phone = combine_phone(request.form.get('phone_code'), phone)
            customer = Customer(company_id=current_user.company_id, name=name, phone=phone, email=email)
            db.session.add(customer)
            log_audit('customer_created', 'Customer', None, f'Customer {name} created.')
            db.session.commit()
            flash('Customer added.', 'success')
            return redirect(url_for('data_manager') if request.referrer and 'data-manager' in request.referrer else url_for('customers'))
        customers = company_query(Customer).order_by(Customer.created_at.desc()).all()
        return render_template('customers.html', customers=customers)

    @app.route('/parts', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff')
    def parts():
        if request.method == 'POST':
            name = request.form['name'].strip()
            compatible_model = request.form.get('compatible_model', '').strip()
            if not name:
                flash('Part name is required.', 'danger')
                return redirect(url_for('parts'))
            if Part.query.filter_by(company_id=current_user.company_id, name=name, compatible_model=compatible_model).first():
                flash('That part already exists.', 'danger')
                return redirect(url_for('parts'))
            part = Part(
                company_id=current_user.company_id,
                name=name,
                compatible_model=compatible_model,
                quantity=int(request.form.get('quantity') or 0),
                selling_price=float(request.form.get('selling_price') or 0),
                low_stock_limit=int(request.form.get('low_stock_limit') or 0),
            )
            db.session.add(part)
            log_audit('part_created', 'Part', None, f'Part {name} created.')
            db.session.commit()
            flash('Part added.', 'success')
            return redirect(url_for('data_manager') if request.referrer and 'data-manager' in request.referrer else url_for('parts'))
        parts = company_query(Part).order_by(Part.created_at.desc()).all()
        return render_template('parts.html', parts=parts)

    @app.route('/customers/import', methods=['POST'])
    @company_required
    @roles_required('Company Admin')
    def import_customers():
        created, skipped, errors = import_customers_csv(request.files.get('file'), current_user.company_id)
        if created:
            flash(f'Imported {created} customers.', 'success')
        if skipped:
            flash(f'Skipped {skipped} duplicate customers.', 'warning')
        for err in errors[:8]:
            flash(err, 'danger')
        return redirect(url_for('data_manager') if request.referrer and 'data-manager' in request.referrer else url_for('customers'))

    @app.route('/parts/import', methods=['POST'])
    @company_required
    @roles_required('Company Admin')
    def import_parts():
        created, skipped, errors = import_parts_csv(request.files.get('file'), current_user.company_id)
        if created:
            flash(f'Imported {created} parts.', 'success')
        if skipped:
            flash(f'Skipped {skipped} duplicate parts.', 'warning')
        for err in errors[:8]:
            flash(err, 'danger')
        return redirect(url_for('data_manager') if request.referrer and 'data-manager' in request.referrer else url_for('parts'))

    @app.route('/import-template/<kind>')
    @company_required
    @roles_required('Company Admin')
    def import_template(kind):
        if kind == 'customers':
            return csv_template_response('customers_template.csv', ['name', 'phone', 'email'])
        if kind == 'parts':
            return csv_template_response('parts_template.csv', ['name', 'compatible_model', 'quantity', 'selling_price', 'low_stock_limit'])
        abort(404)

    @app.route('/repairs', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def repairs():
        if request.method == 'POST':
            customer_name = request.form['customer_name'].strip()
            customer_phone = combine_phone(
                request.form.get('customer_phone_code'),
                request.form.get('customer_phone', '')
            )
            customer_email = request.form.get('customer_email', '').strip()
            items = parse_repair_items_from_form(request.form)
            if not customer_name or not items:
                flash('Customer and at least one device are required.', 'danger')
                return redirect(url_for('repairs'))
            customer = company_query(Customer).filter_by(name=customer_name, phone=customer_phone).first()
            if not customer:
                customer = Customer(company_id=current_user.company_id, name=customer_name, phone=customer_phone or '', email=customer_email)
                db.session.add(customer)
                db.session.flush()
            first = items[0]
            repair = RepairJob(
                company_id=current_user.company_id,
                repair_code=generate_repair_code(current_user.company_id),
                invoice_number=generate_invoice_number(current_user.company_id),
                customer_id=customer.id,
                device_brand=first.get('brand', ''),
                device_model=first.get('model', ''),
                serial_number=first.get('serial', ''),
                imei=first.get('imei', ''),
                issue_description=first.get('issue', ''),
                status=request.form.get('status', 'Received'),
                estimated_cost=sum(safe_float(item.get('amount'), 0) for item in items),
                deposit_amount=float(request.form.get('deposit_amount') or 0),
                due_date=parse_date_field(request.form.get('due_date')),
                customer_note=request.form.get('customer_note', '').strip(),
                items_json=json.dumps(items),
            )
            db.session.add(repair)
            log_audit('repair_created', 'RepairJob', None, f'Repair {repair.repair_code} created for {customer.name}.')
            db.session.commit()
            flash('Repair job created.', 'success')
            return redirect(url_for('repairs'))
        repairs = company_query(RepairJob).order_by(RepairJob.created_at.desc()).all()
        return render_template('repairs.html', repairs=repairs, statuses=get_statuses())

    @app.route('/repairs/<int:repair_id>', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def repair_detail(repair_id):
        repair = company_object_or_404(RepairJob, repair_id)
        if request.method == 'POST':
            action = request.form.get('action')
            if action == 'update_status':
                old_status = repair.status
                repair.status = request.form['status']
                update = RepairUpdate(company_id=current_user.company_id, repair_job_id=repair.id, user_id=current_user.id,
                                      old_status=old_status, new_status=repair.status, note=request.form.get('note', '').strip())
                db.session.add(update)
                log_audit('status_changed', 'RepairJob', repair.id, f'{repair.repair_code}: {old_status or '-'} -> {repair.status}.')
                db.session.commit()
                flash('Repair status updated.', 'success')
            elif action == 'add_payment':
                payment_id = request.form.get('payment_id')
                amount = safe_float(request.form.get('amount'), 0)
                method = (request.form.get('method', 'Cash') or 'Cash').strip()
                if amount <= 0:
                    flash('Payment amount must be greater than zero.', 'danger')
                    return redirect(url_for('repair_detail', repair_id=repair.id))
                if payment_id:
                    payment = company_object_or_404(Payment, int(payment_id))
                    payment.amount = amount
                    payment.method = method
                    log_audit('payment_updated', 'Payment', payment.id, f'{repair.repair_code}: payment updated to €{payment.amount:.2f} via {payment.method}.')
                    flash('Payment updated.', 'success')
                else:
                    payment = Payment(company_id=current_user.company_id, repair_job_id=repair.id, amount=amount, method=method)
                    db.session.add(payment)
                    log_audit('payment_added', 'RepairJob', repair.id, f'{repair.repair_code}: payment €{payment.amount:.2f} via {payment.method}.')
                    flash('Payment recorded.', 'success')
                db.session.commit()
            elif action == 'use_part' and current_user.role in ('Company Admin', 'Staff', 'Technician'):
                part_id = request.form.get('part_id')
                if not part_id:
                    flash('Please choose a part.', 'danger')
                else:
                    part = company_object_or_404(Part, int(part_id))
                    qty = int(request.form.get('quantity') or 1)
                    if qty <= 0 or part.quantity < qty:
                        flash('Not enough stock for that part.', 'danger')
                    else:
                        part.quantity -= qty
                        used = RepairPartUsed(company_id=current_user.company_id, repair_job_id=repair.id, part_id=part.id, quantity=qty, price_used=part.selling_price)
                        db.session.add(used)
                        log_audit('part_used', 'RepairJob', repair.id, f'{repair.repair_code}: used {part.name} x{qty}.')
                        db.session.commit()
                        flash('Part used and stock reduced.', 'success')
            elif action == 'add_internal_note':
                note_text = request.form.get('internal_note', '').strip()
                if note_text:
                    db.session.add(InternalNote(company_id=current_user.company_id, repair_job_id=repair.id, user_id=current_user.id, note=note_text))
                    log_audit('internal_note_added', 'RepairJob', repair.id, f'Internal note added to {repair.repair_code}.')
                    db.session.commit()
                    flash('Internal note added.', 'success')
            elif action == 'update_customer_note':
                repair.customer_note = request.form.get('customer_note', '').strip()
                log_audit('customer_note_updated', 'RepairJob', repair.id, f'Customer note updated for {repair.repair_code}.')
                db.session.commit()
                flash('Customer note updated.', 'success')
            return redirect(url_for('repair_detail', repair_id=repair.id))
        parts = company_query(Part).order_by(Part.name.asc()).all()
        recorded_payments = sum(p.amount for p in repair.payments)
        total_paid = (repair.deposit_amount or 0) + recorded_payments
        items = repair_items(repair)
        total_parts = sum((p.price_used or 0) * p.quantity for p in repair.parts_used)
        subtotal = max(repair_items_total(repair), 0) + total_parts
        balance = max(subtotal - total_paid, 0)
        timeline = []
        timeline += [{'when': r.created_at, 'kind': 'status', 'label': f'{r.old_status or "-"} → {r.new_status}', 'by': r.user.name, 'note': r.note or ''} for r in repair.updates]
        timeline += [{'when': p.created_at, 'kind': 'payment', 'label': f'Payment: € {p.amount:.2f}', 'by': '', 'note': p.method} for p in repair.payments]
        timeline += [{'when': u.created_at, 'kind': 'part', 'label': f'Part used: {u.part.name} x{u.quantity}', 'by': '', 'note': f'€ {u.price_used:.2f}'} for u in repair.parts_used]
        timeline += [{'when': n.created_at, 'kind': 'internal', 'label': 'Internal note', 'by': n.user.name, 'note': n.note} for n in repair.internal_notes]
        timeline = sorted(timeline, key=lambda x: x['when'], reverse=True)
        payment_to_edit = None
        edit_payment_id = request.args.get('edit_payment_id', type=int)
        if edit_payment_id:
            candidate = Payment.query.filter_by(id=edit_payment_id, company_id=current_user.company_id, repair_job_id=repair.id).first()
            if candidate:
                payment_to_edit = candidate
        return render_template('repair_detail.html', repair=repair, parts=parts, total_paid=total_paid, subtotal=subtotal, balance=balance, timeline=timeline, payment_methods=get_payment_methods(), statuses=get_statuses(), payment_to_edit=payment_to_edit)



    @app.route('/repairs/<int:repair_id>/notify-customer')
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def notify_customer(repair_id):
        repair = company_object_or_404(RepairJob, repair_id)
        if not repair.customer or not normalize_phone(repair.customer.phone):
            flash('Customer phone number is missing.', 'danger')
            return redirect(url_for('repair_detail', repair_id=repair.id))
        wa_url = build_customer_whatsapp_url(repair)
        log_audit('customer_whatsapp_opened', 'RepairJob', repair.id, f'WhatsApp notification opened for {repair.repair_code}.')
        db.session.commit()
        return redirect(wa_url)

    @app.route('/payments/<int:payment_id>/edit', methods=['POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def edit_payment(payment_id):
        payment = company_object_or_404(Payment, payment_id)
        amount = safe_float(request.form.get('amount'), payment.amount)
        method = (request.form.get('method') or payment.method).strip()
        if amount <= 0:
            flash('Payment amount must be greater than zero.', 'danger')
            return redirect(url_for('repair_detail', repair_id=payment.repair_job_id))
        payment.amount = amount
        payment.method = method or payment.method
        log_audit('payment_updated', 'Payment', payment.id, f'Payment for {payment.repair_job.repair_code} updated to €{payment.amount:.2f} via {payment.method}.')
        db.session.commit()
        flash('Payment updated.', 'success')
        return redirect(url_for('repair_detail', repair_id=payment.repair_job_id))

    @app.route('/payments/<int:payment_id>/delete', methods=['POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def delete_payment(payment_id):
        payment = company_object_or_404(Payment, payment_id)
        repair_id = payment.repair_job_id
        detail = f'Payment for {payment.repair_job.repair_code} deleted (€{payment.amount:.2f} via {payment.method}).'
        db.session.delete(payment)
        log_audit('payment_deleted', 'Payment', payment.id, detail)
        db.session.commit()
        flash('Payment deleted.', 'success')
        return redirect(url_for('repair_detail', repair_id=repair_id))

    @app.route('/repairs/<int:repair_id>/invoice')
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def print_invoice(repair_id):
        repair = company_object_or_404(RepairJob, repair_id)
        settings = AppSetting.query.first()
        size = request.args.get('size', '80mm')
        lang = request.args.get('lang', current_user.company.default_language or 'en')
        recorded_payments = sum(p.amount for p in repair.payments)
        total_paid = (repair.deposit_amount or 0) + recorded_payments
        items = repair_items(repair)
        total_parts = sum((p.price_used or 0) * p.quantity for p in repair.parts_used)
        subtotal = max(repair_items_total(repair), 0) + total_parts
        balance = max(subtotal - total_paid, 0)
        size_class = {'80mm': 'size-80mm', '4x6': 'size-4x6', 'a4': 'size-a4'}.get(size, 'size-4x6')
        terms = getattr(repair.company, f'invoice_terms_{lang}', None) or repair.company.invoice_terms_en
        notice = getattr(repair.company, f'storage_notice_{lang}', None) or repair.company.storage_notice_en
        qr_image = qr_data_uri(repair.repair_code)
        barcode_image = barcode_svg_uri(repair.repair_code)
        public_url = request.url_root.rstrip('/') + url_for('public_status_portal') + f'?repair={repair.repair_code}'
        return render_template(
            'print_invoice.html', repair=repair, items=items, company=repair.company, size=size, size_class=size_class,
            total_paid=total_paid, subtotal=subtotal, balance=balance, invoice_terms_text=terms,
            storage_notice_text=notice, invoice_lang=lang, qr_image=qr_image, barcode_image=barcode_image,
            public_url=public_url, invoice_title=(settings.invoice_title if settings else 'Repair Invoice'),
            invoice_line_label=(settings.invoice_line_label if settings else 'Problem Description'),
            qr_caption=(settings.qr_caption if settings else 'Scan for status'),
            show_public_status_url=(settings.show_public_status_url if settings else False),
            invoice_id_label=t('invoice_repair_id'), repair_code_label=t('repair_code'), customer_label=t('customer'),
            phone_label=t('phone'), device_label=t('device'), serial_label=t('serial_number'), status_label_text=t('status'),
            description_label=t('description'), qty_label=t('qty'), amount_label=t('amount'), subtotal_label=t('subtotal'),
            deposit_label=t('deposit'), total_paid_label=t('total_paid'), balance_due_label=t('balance_due')
        )





    @app.route('/repairs/finish-all', methods=['POST'])
    @company_required
    @roles_required('Company Admin', 'Staff')
    def finish_all_repairs():
        count = RepairJob.query.filter(
            RepairJob.company_id == current_user.company_id,
            RepairJob.is_archived.is_(False),
            RepairJob.status != 'Delivered'
        ).update({'status': 'Delivered'}, synchronize_session=False)
        log_audit('repairs_finish_all', 'RepairJob', None, f'{count} repairs marked as Delivered.')
        db.session.commit()
        flash(f'{count} repairs marked as Delivered.', 'success')
        return redirect(url_for('repairs'))

    @app.route('/repairs/<int:repair_id>/edit', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def edit_repair(repair_id):
        repair = company_object_or_404(RepairJob, repair_id)
        if request.method == 'POST':
            repair.customer.name = request.form.get('customer_name', repair.customer.name).strip() or repair.customer.name
            phone_local = request.form.get('customer_phone')
            phone_code = request.form.get('customer_phone_code')
            if (phone_local or '').strip():
                repair.customer.phone = combine_phone(phone_code, phone_local)
            else:
                repair.customer.phone = repair.customer.phone or ''
            repair.customer.email = request.form.get('customer_email', repair.customer.email or '').strip()
            items = parse_repair_items_from_form(request.form)
            if not items:
                flash('At least one device is required.', 'danger')
                return redirect(url_for('edit_repair', repair_id=repair.id))
            first = items[0]
            repair.device_brand = first.get('brand', repair.device_brand).strip() or repair.device_brand
            repair.device_model = first.get('model', repair.device_model).strip() or repair.device_model
            repair.imei = first.get('imei', repair.imei or '').strip()
            repair.serial_number = first.get('serial', repair.serial_number or '').strip()
            repair.due_date = parse_date_field(request.form.get('due_date'))
            repair.issue_description = first.get('issue', repair.issue_description).strip() or repair.issue_description
            repair.status = request.form.get('status', repair.status)
            repair.estimated_cost = sum(safe_float(item.get('amount'), 0) for item in items)
            repair.deposit_amount = safe_float(request.form.get('deposit_amount'), repair.deposit_amount or 0)
            repair.customer_note = request.form.get('customer_note', repair.customer_note or '').strip()
            repair.items_json = json.dumps(items)
            log_audit('repair_updated', 'RepairJob', repair.id, f'Repair {repair.repair_code} updated.')
            db.session.commit()
            flash('Repair updated.', 'success')
            return redirect(url_for('repair_detail', repair_id=repair.id))
        return render_template('repair_edit.html', repair=repair, statuses=get_statuses(), items=repair_items(repair))

    @app.route('/repairs/<int:repair_id>/delete', methods=['POST'])
    @company_required
    @roles_required('Company Admin', 'Staff')
    def delete_repair(repair_id):
        repair = company_object_or_404(RepairJob, repair_id)
        log_audit('repair_deleted', 'RepairJob', repair.id, f'Repair {repair.repair_code} deleted.')
        for note in repair.internal_notes:
            db.session.delete(note)
        for update in repair.updates:
            db.session.delete(update)
        for used in repair.parts_used:
            db.session.delete(used)
        for payment in repair.payments:
            db.session.delete(payment)
        db.session.delete(repair)
        db.session.commit()
        flash('Repair deleted.', 'success')
        return redirect(url_for('repairs'))

    @app.route('/buying', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def buying():
        buying_statuses = get_gsm_statuses()
        if request.method == 'POST':
            customer_name = request.form.get('customer_name', '').strip()
            customer_phone = combine_phone(request.form.get('customer_phone_code'), request.form.get('customer_phone', ''))
            customer_email = request.form.get('customer_email', '').strip()
            note = request.form.get('issue_description', '').strip()
            items = parse_gsm_items_from_form(request.form)
            if not customer_name:
                flash('Customer name is required.', 'danger')
                return redirect(url_for('buying'))
            if not items:
                flash('Please add at least one GSM item.', 'danger')
                return redirect(url_for('buying'))
            customer = company_query(Customer).filter_by(name=customer_name, phone=customer_phone).first()
            if not customer:
                customer = Customer(company_id=current_user.company_id, name=customer_name, phone=customer_phone or '', email=customer_email)
                db.session.add(customer)
                db.session.flush()
            first = items[0]
            ticket = BuyingTicket(
                company_id=current_user.company_id,
                ticket_code=generate_ticket_code(current_user.company_id),
                invoice_number=generate_buying_invoice_number(current_user.company_id),
                customer_id=customer.id,
                device_brand=first.get('brand',''),
                device_model=first.get('model',''),
                serial_number=first.get('serial',''),
                imei=first.get('imei',''),
                battery_percentage=first.get('battery',''),
                storage_gb=first.get('gb',''),
                issue_description=note,
                status=request.form.get('status', get_gsm_statuses()[0]),
                estimated_cost=sum((item.get('qty',1) or 1) * safe_float(item.get('amount'),0) for item in items),
                deposit_amount=0,
                due_date=parse_date_field(request.form.get('due_date')),
                customer_note='',
                items_json=json.dumps(items),
            )
            db.session.add(ticket)
            log_audit('buying_created', 'BuyingTicket', None, f'GSM ticket {ticket.ticket_code} created for {customer.name}.')
            db.session.commit()
            flash('Ticket created.', 'success')
            return redirect(url_for('buying'))
        # Normalize old GSM statuses to current labels
        old_map = {'Buying': get_gsm_statuses()[0], 'Selling': get_gsm_statuses()[1]}
        changed = False
        for tkt in company_query(BuyingTicket).all():
            if tkt.status in old_map:
                tkt.status = old_map[tkt.status]
                changed = True
        if changed:
            db.session.commit()
        tickets = company_query(BuyingTicket).order_by(BuyingTicket.created_at.desc()).all()
        return render_template('buying.html', tickets=tickets, statuses=buying_statuses, items=[{'brand':'','model':'','gb':'','imei':'','battery':'','serial':'','qty':1,'amount':0}])

    @app.route('/buying/<int:ticket_id>', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def buying_detail(ticket_id):
        ticket = company_object_or_404(BuyingTicket, ticket_id)
        buying_statuses = get_gsm_statuses()
        if request.method == 'POST':
            action = request.form.get('action')
            if action == 'update_status':
                ticket.status = request.form.get('status', get_gsm_statuses()[0])
                log_audit('buying_status_changed', 'BuyingTicket', ticket.id, f'{ticket.ticket_code}: status -> {ticket.status}.')
                db.session.commit()
                flash('Ticket status updated.', 'success')
            elif action == 'add_payment':
                payment_id = request.form.get('payment_id')
                amount = safe_float(request.form.get('amount'), 0)
                method = (request.form.get('method', 'Cash') or 'Cash').strip()
                if amount <= 0:
                    flash('Payment amount must be greater than zero.', 'danger')
                    return redirect(url_for('buying_detail', ticket_id=ticket.id))
                if payment_id:
                    payment = company_object_or_404(BuyingPayment, int(payment_id))
                    payment.amount = amount
                    payment.method = method
                    log_audit('buying_payment_updated', 'BuyingPayment', payment.id, f'{ticket.ticket_code}: payment updated to €{payment.amount:.2f} via {payment.method}.')
                    flash('Payment updated.', 'success')
                else:
                    payment = BuyingPayment(company_id=current_user.company_id, buying_ticket_id=ticket.id, amount=amount, method=method)
                    db.session.add(payment)
                    log_audit('buying_payment_added', 'BuyingTicket', ticket.id, f'{ticket.ticket_code}: payment €{payment.amount:.2f} via {payment.method}.')
                    flash('Payment recorded.', 'success')
                db.session.commit()
            return redirect(url_for('buying_detail', ticket_id=ticket.id))
        total = ticket_total(ticket)
        return render_template('buying_detail.html', ticket=ticket, total=total, items=ticket_items(ticket), payment_methods=get_payment_methods(), statuses=buying_statuses)

    @app.route('/buying/<int:ticket_id>/edit', methods=['GET', 'POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def edit_buying(ticket_id):
        ticket = company_object_or_404(BuyingTicket, ticket_id)
        buying_statuses = get_gsm_statuses()
        if request.method == 'POST':
            customer = ticket.customer
            customer.name = request.form.get('customer_name', '').strip()
            new_phone = combine_phone(request.form.get('customer_phone_code'), request.form.get('customer_phone', ''))
            if new_phone:
                customer.phone = new_phone
            customer.email = request.form.get('customer_email', '').strip()
            items = parse_gsm_items_from_form(request.form)
            if not items:
                flash('Please add at least one GSM item.', 'danger')
                return redirect(url_for('edit_buying', ticket_id=ticket.id))
            first = items[0]
            ticket.device_brand = first.get('brand', '')
            ticket.device_model = first.get('model', '')
            ticket.serial_number = first.get('serial', '')
            ticket.imei = first.get('imei', '')
            ticket.battery_percentage = first.get('battery', '')
            ticket.storage_gb = first.get('gb', '')
            ticket.issue_description = request.form.get('issue_description', '').strip()
            ticket.status = request.form.get('status', get_gsm_statuses()[0])
            ticket.estimated_cost = sum((item.get('qty', 1) or 1) * safe_float(item.get('amount'), 0) for item in items)
            ticket.deposit_amount = 0
            ticket.due_date = parse_date_field(request.form.get('due_date'))
            ticket.customer_note = ''
            ticket.items_json = json.dumps(items)
            log_audit('buying_updated', 'BuyingTicket', ticket.id, f'{ticket.ticket_code} updated.')
            db.session.commit()
            flash('Ticket updated.', 'success')
            return redirect(url_for('buying_detail', ticket_id=ticket.id))
        return render_template('buying_edit.html', ticket=ticket, statuses=buying_statuses, items=ticket_items(ticket))

    @app.route('/buying/<int:ticket_id>/delete', methods=['POST'])
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def delete_buying(ticket_id):
        ticket = company_object_or_404(BuyingTicket, ticket_id)
        BuyingPayment.query.filter_by(buying_ticket_id=ticket.id).delete()
        db.session.delete(ticket)
        log_audit('buying_deleted', 'BuyingTicket', ticket.id, f'{ticket.ticket_code} deleted.')
        db.session.commit()
        flash('Ticket deleted.', 'success')
        return redirect(url_for('buying'))

    @app.route('/buying/<int:ticket_id>/invoice')
    @company_required
    @roles_required('Company Admin', 'Staff', 'Technician')
    def print_buying_invoice(ticket_id):
        ticket = company_object_or_404(BuyingTicket, ticket_id)
        size = request.args.get('size', '80mm')
        lang = request.args.get('lang', current_user.company.default_language or 'en')
        items = ticket_items(ticket)
        total = ticket_total(ticket)
        total_paid = sum(p.amount for p in ticket.payments)
        subtotal = total
        balance = max(total - total_paid, 0)
        size_class = {'80mm': 'size-80mm', '4x6': 'size-4x6', 'a4': 'size-a4'}.get(size, 'size-4x6')
        terms = getattr(ticket.company, f'gsm_terms_{lang}', None) or getattr(ticket.company, 'gsm_terms_en', '') or ''
        notice = getattr(ticket.company, f'gsm_notice_{lang}', None) or getattr(ticket.company, 'gsm_notice_en', '') or ''
        qr_image = qr_data_uri(ticket.ticket_code)
        barcode_image = barcode_svg_uri(ticket.ticket_code)
        public_url = ''
        return render_template('print_buying_invoice.html', ticket=ticket, items=items, total=total, company=ticket.company, size=size, size_class=size_class, total_paid=total_paid, subtotal=subtotal, balance=balance, invoice_terms_text=terms, storage_notice_text=notice, invoice_lang=lang, qr_image=qr_image, barcode_image=barcode_image, public_url=public_url, invoice_title=(getattr(ticket.company, 'gsm_invoice_title', None) or 'Ticket'))


    @app.route('/reports')
    @company_required
    @roles_required('Company Admin', 'Staff')
    def reports():
        date_from = parse_date_field(request.args.get('date_from'))
        date_to = parse_date_field(request.args.get('date_to'))
        repairs_q = company_query(RepairJob)
        payments_q = Payment.query.filter_by(company_id=current_user.company_id)
        parts_q = RepairPartUsed.query.filter_by(company_id=current_user.company_id)
        if date_from:
            repairs_q = repairs_q.filter(RepairJob.created_at >= datetime.combine(date_from, datetime.min.time()))
            payments_q = payments_q.filter(Payment.created_at >= datetime.combine(date_from, datetime.min.time()))
            parts_q = parts_q.filter(RepairPartUsed.created_at >= datetime.combine(date_from, datetime.min.time()))
        if date_to:
            end_dt = datetime.combine(date_to + timedelta(days=1), datetime.min.time())
            repairs_q = repairs_q.filter(RepairJob.created_at < end_dt)
            payments_q = payments_q.filter(Payment.created_at < end_dt)
            parts_q = parts_q.filter(RepairPartUsed.created_at < end_dt)
        repairs = repairs_q.all()
        payments = payments_q.all()
        part_rows = parts_q.all()
        status_counts = dict(db.session.query(RepairJob.status, func.count(RepairJob.id)).filter_by(company_id=current_user.company_id, is_archived=False).group_by(RepairJob.status).all())
        revenue = sum(p.amount for p in payments)
        deposit_total = sum(r.deposit_amount or 0 for r in repairs)
        part_usage = {}
        for row in part_rows:
            part_usage.setdefault(row.part.name, 0)
            part_usage[row.part.name] += row.quantity
        top_parts = sorted(part_usage.items(), key=lambda x: x[1], reverse=True)[:10]
        return render_template('reports.html', repairs=repairs, payments=payments, revenue=revenue, deposit_total=deposit_total, status_counts=status_counts, top_parts=top_parts, date_from=date_from.isoformat() if date_from else '', date_to=date_to.isoformat() if date_to else '')

    @app.route('/audit-log')
    @login_required
    def audit_log_page():
        if current_user.is_super_admin:
            logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
        else:
            logs = AuditLog.query.filter_by(company_id=current_user.company_id).order_by(AuditLog.created_at.desc()).limit(200).all()
        return render_template('audit_log.html', logs=logs)

    @app.route('/support', methods=['GET'])
    def public_status_portal():
        repair_code = request.args.get('repair', '').strip()
        phone = request.args.get('phone', '').strip()
        repair = None
        if repair_code:
            repair = RepairJob.query.filter_by(repair_code=repair_code).first()
        elif phone:
            repair = RepairJob.query.join(Customer).filter(Customer.phone == phone, RepairJob.is_archived.is_(False)).order_by(RepairJob.created_at.desc()).first()
        return render_template('public_status.html', repair=repair, repair_code=repair_code, phone=phone)


    @app.errorhandler(403)
    def forbidden(_error):
        return render_template('error.html', code=403, title='Access denied', message='You do not have permission to access this page.'), 403

    @app.errorhandler(404)
    def not_found(_error):
        return render_template('error.html', code=404, title='Page not found', message='The page you requested could not be found.'), 404

    @app.errorhandler(500)
    def server_error(_error):
        db.session.rollback()
        return render_template('error.html', code=500, title='Server error', message='Something went wrong on the server. Please try again.'), 500

    return app